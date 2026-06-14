# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')

lines = []
lines.append("import os")
lines.append("import re")
lines.append("import json")
lines.append("import time")
lines.append("import random")
lines.append("import threading")
lines.append("from datetime import datetime")
lines.append("from urllib.parse import urlparse")
lines.append("")
lines.append("import redis as redis_module")
lines.append("import requests")
lines.append("import pandas as pd")
lines.append("from openpyxl import Workbook, load_workbook")
lines.append("from rapidfuzz import fuzz")
lines.append("")
lines.append('PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))')
lines.append('DATA_DIR = os.path.join(PROJECT_DIR, "Data", "\u5546\u54c1\u5bfc\u51fa")')
lines.append('CURRENCY_CONFIG_PATH = os.path.join(PROJECT_DIR, "currency_config.json")')
lines.append('PROXIES_FILE = os.path.join(PROJECT_DIR, "proxies.txt")')
lines.append('BESTPROXY_TOKENS_FILE = os.path.join(PROJECT_DIR, "bestproxy_tokens.txt")')
lines.append("")
lines.append("PER_CATEGORY_MIN = 20")
lines.append("MAX_TOTAL = 60000")
lines.append("FLUSH_INTERVAL = 500")
lines.append("MATCH_THRESHOLD_PHASE1 = 75")
lines.append("MATCH_THRESHOLD_PHASE2 = 60")
lines.append("MIN_SHOPIFY_COUNT = 200")
lines.append("PAGE_SLEEP = (0.5, 1.5)")
lines.append("SITE_COOLDOWN = (1.0, 2.0)")
lines.append("MAX_PAGES = 50")
lines.append("MAX_EMPTY_PAGES = 3")
lines.append("MIN_PRICE = 1.0")
lines.append("")
lines.append('EXPORT_COLUMNS = ["SKU", "\u6807\u9898", "\u63cf\u8ff0", "\u5b50\u63cf\u8ff0", "\u56fe\u7247", "\u539f\u4ef7", "\u6298\u6263\u4ef7", "\u53d8\u4f53\u540d", "\u53d8\u4f53\u503c", "\u5206\u7c7b"]')
lines.append("")

# I'll use a different approach - write the whole thing from a Python script file
script = r'''
import os

content = """import os
import re
import json
import time
import random
import threading
from datetime import datetime
from urllib.parse import urlparse

import redis as redis_module
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from rapidfuzz import fuzz

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(PROJECT_DIR, "Data", "\u5546\u54c1\u5bfc\u51fa")
CURRENCY_CONFIG_PATH = os.path.join(PROJECT_DIR, "currency_config.json")
PROXIES_FILE = os.path.join(PROJECT_DIR, "proxies.txt")
BESTPROXY_TOKENS_FILE = os.path.join(PROJECT_DIR, "bestproxy_tokens.txt")

PER_CATEGORY_MIN = 20
MAX_TOTAL = 60000
FLUSH_INTERVAL = 500
MATCH_THRESHOLD_PHASE1 = 75
MATCH_THRESHOLD_PHASE2 = 60
MIN_SHOPIFY_COUNT = 200
PAGE_SLEEP = (0.5, 1.5)
SITE_COOLDOWN = (1.0, 2.0)
MAX_PAGES = 50
MAX_EMPTY_PAGES = 3
MIN_PRICE = 1.0

EXPORT_COLUMNS = ["SKU", "\u6807\u9898", "\u63cf\u8ff0", "\u5b50\u63cf\u8ff0", "\u56fe\u7247", "\u539f\u4ef7", "\u6298\u6263\u4ef7", "\u53d8\u4f53\u540d", "\u53d8\u4f53\u503c", "\u5206\u7c7b"]

_REDIS_CLIENT = None
_REDIS_LOCK = threading.Lock()


def _get_redis():
    global _REDIS_CLIENT
    if _REDIS_CLIENT is None:
        with _REDIS_LOCK:
            if _REDIS_CLIENT is None:
                try:
                    _REDIS_CLIENT = redis_module.Redis(
                        host="localhost", port=6379, db=0,
                        socket_connect_timeout=3, socket_timeout=5,
                        decode_responses=True,
                    )
                    _REDIS_CLIENT.ping()
                except Exception:
                    _REDIS_CLIENT = None
    return _REDIS_CLIENT


def _redis_key(task_id, suffix):
    return f"structured_crawl:{task_id}:{suffix}"


def _load_currency_map():
    try:
        with open(CURRENCY_CONFIG_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        result = {}
        for nation, info in data.items():
            rate = info.get("exchange_rate_usd")
            if rate:
                code = info.get("currency_code", nation).upper()
                try:
                    result[code] = float(rate)
                except (ValueError, TypeError):
                    pass
        return result
    except Exception:
        return {}


def _convert_price(price_str, rate):
    if not price_str:
        return ""
    try:
        p = float(str(price_str).replace(",", "").strip())
        if p <= 0:
            return ""
        if rate and rate > 0:
            return f"{p / rate:.2f}"
        return f"{p:.2f}"
    except (ValueError, TypeError):
        return ""


def _extract_images(product):
    images = product.get("images") or []
    if not images:
        return ""
    img = images[0]
    if isinstance(img, dict):
        src = img.get("src") or img.get("src") or ""
    elif isinstance(img, str):
        src = img
    else:
        return ""
    if "?" in src:
        src = src.split("?")[0]
    return src


def _extract_variant_info(variants, options):
    if not variants:
        return "", ""
    variant = variants[0]
    sku = str(variant.get("sku") or "").strip() if isinstance(variant, dict) else ""
    option_names = []
    for opt in (options or []):
        if isinstance(opt, dict):
            name = str(opt.get("name") or "").strip()
            if name:
                option_names.append(name)
    values_list = []
    for v in (variants or []):
        if isinstance(v, dict):
            vals = v.get("title") or ""
            if vals and vals.lower() != "default title":
                values_list.append(str(vals).strip())
    if option_names or values_list:
        variant_str = "|||".join(
            f"{n}^{v}" for n, v in zip(option_names or [""], values_list or [""])
        )
    else:
        variant_str = ""
    return sku, variant_str


def _extract_prices(variants):
    if not variants:
        return "", ""
    variant = variants[0] if isinstance(variants, list) else variants
    if not isinstance(variant, dict):
        return "", ""
    compare_at = variant.get("compare_at_price") or ""
    price = variant.get("price") or ""
    return str(compare_at).strip(), str(price).strip()


def _product_unique_key(domain, product_id, title, image):
    if product_id:
        return f"{domain}|||id|||{product_id}"
    return f"{domain}|||fallback|||{title}|||{image}"


def _is_non_english_products(products):
    import re as _re
    sample_texts = []
    for p in products[:20]:
        if not isinstance(p, dict):
            continue
        title = str(p.get("title") or "")
        desc = _re.sub(r"<[^>]+>", "", str(p.get("body_html") or ""))
        combined = f"{title} {desc}".strip()
        if len(combined) >= 20:
            sample_texts.append(combined)
    if not sample_texts:
        return False
    sample = " ".join(sample_texts)
    latin = len(_re.findall(r"[a-zA-Z]", sample))
    total_chars = len(_re.findall(r"[a-zA-Z0-9]", sample))
    if total_chars < 30:
        return False
    non_latin = len(_re.findall(
        r"[\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff\u0400-\u04ff\u0600-\u06ff\u0e00-\u0e7f\u0e80-\u0eff"
        r"\uac00-\ud7af\u0590-\u05ff\u0700-\u074f\u0750-\u077f\u0f00-\u0fff\u0a80-\u0aff"
        r"\u0b00-\u0b7f\u0b80-\u0bff\u0c00-\u0c7f\u0c80-\u0cff\u0d00-\u0d7f\u0d80-\u0dff]",
        sample,
    ))
    if non_latin > 0:
        return True
    european_accent = len(_re.findall(r"[\u00e0\u00e1\u00e2\u00e3\u00e4\u00e5\u00e6\u00e7\u00e8\u00e9\u00ea\u00eb\u00ec\u00ed\u00ee\u00ef\u00f0\u00f1\u00f2\u00f3\u00f4\u00f5\u00f6\u00f8\u00f9\u00fa\u00fb\u00fc\u00fd\u00fe]", sample.lower()))
    if latin > 0 and european_accent / latin > 0.3:
        return True
    if latin < 0.6 * total_chars and total_chars >= 30:
        return True
    return False


def _search_shopify_sites(
    keyword,
    api_mode="5",
    max_results=50,
    min_count=MIN_SHOPIFY_COUNT,
    progress_callback=None,
    stop_callback=None,
):
    from data_scraper import run_scraper_job as search_job

    if progress_callback:
        progress_callback(f"[\u641c\u7d22] \u5173\u952e\u8bcd: {keyword}")

    bestproxy_auth = None
    bestproxy_tokens_path = BESTPROXY_TOKENS_FILE
    if os.path.exists(bestproxy_tokens_path):
        try:
            with open(bestproxy_tokens_path, "r") as f:
                tokens = [t.strip() for t in f if t.strip()]
                if tokens:
                    bestproxy_auth = tokens[0]
        except Exception:
            pass

    result = search_job(
        keywords=[f"{keyword} inurl:collections/all"],
        max_results=max_results,
        min_product_count=min_count,
        api_mode=api_mode,
        api_key=None,
        bestproxy_auth=bestproxy_auth,
        save_mode="excel",
        category=f"_struct_{keyword[:20]}",
        mongo_collection=None,
        progress_callback=progress_callback,
        stop_callback=stop_callback,
    )

    urls = []
    for item in result.get("results", []):
        if len(item) >= 3 and item[1] == "Shopify":
            count = int(item[2]) if str(item[2]).isdigit() else 0
            if count >= min_count:
                urls.append((item[0], count))
    return urls


def _crawl_and_match(
    url,
    secondary_cat,
    primary_cat,
    full_label,
    currency_map,
    seen_unique_keys,
    redis_conn,
    redis_key_list,
    redis_key_count,
    buffer,
    buffer_lock,
    excel_path,
    progress_callback=None,
    stop_callback=None,
):
    domain = urlparse(url).netloc if "://" in url else urlparse(f"https://{url}").netloc
    if progress_callback:
        progress_callback(f"  [\u722c\u53d6] {domain}")

    if not url.startswith(("http://", "https://")):
        url = f"https://{url}"
    url = url.rstrip("/")

    try:
        meta_resp = requests.get(f"{url}/meta.json", timeout=12)
        if meta_resp.status_code != 200:
            return
        meta = meta_resp.json()
        currency = (meta.get("currency") or "USD").upper()
        rate = currency_map.get(currency)
        if rate is None:
            if progress_callback:
                progress_callback(f"  [\u8df3\u8fc7] {domain} \u65e0\u6c47\u7387\u914d\u7f6e: {currency}")
            return
    except Exception:
        return

    page = 1
    empty_pages = 0
    page_all_matched = 0

    while page <= MAX_PAGES and empty_pages < MAX_EMPTY_PAGES:
        if stop_callback and stop_callback():
            return

        try:
            resp = requests.get(
                f"{url}/products.json?limit=250&page={page}",
                timeout=20,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                },
            )
            if resp.status_code != 200:
                break
            data = resp.json()
            products = data.get("products", [])
        except Exception:
            break

        if not products:
            empty_pages += 1
            page += 1
            time.sleep(random.uniform(*PAGE_SLEEP))
            continue

        empty_pages = 0

        if page == 1:
            if _is_non_english_products(products):
                if progress_callback:
                    progress_callback(f"  [\u8df3\u8fc7] {domain} \u975e\u82f1\u6587\u7ad9\u70b9")
                return

        for product in products:
            if not isinstance(product, dict):
                continue

            title = str(product.get("title") or "").strip()
            body = str(product.get("body_html") or "").strip()
            if not title or not body:
                continue

            images = product.get("images", [])
            variants = product.get("variants", [])
            options = product.get("options", [])
            product_type = str(product.get("product_type") or "").strip()

            image = _extract_images(product)
            sku, variant_str = _extract_variant_info(variants, options)
            compare_at, price = _extract_prices(variants)
            original_price = _convert_price(compare_at, rate)
            discount_price = _convert_price(price, rate)
            price_val = discount_price if discount_price else original_price
            if not price_val or float(price_val) < MIN_PRICE:
                continue

            product_id = str(product.get("id") or "")
            unique_key = _product_unique_key(domain, product_id, title, image)
            if unique_key in seen_unique_keys:
                continue
            seen_unique_keys.add(unique_key)

            cat_value = (product_type or "").strip()
            title_value = title

            cat_score = fuzz.token_set_ratio(secondary_cat, cat_value)
            title_score = fuzz.token_set_ratio(secondary_cat, title_value)
            if cat_score < MATCH_THRESHOLD_PHASE1 and title_score < MATCH_THRESHOLD_PHASE1:
                continue

            page_all_matched += 1
            row = {
                "SKU": sku,
                "\u6807\u9898": title,
                "\u63cf\u8ff0": body,
                "\u5b50\u63cf\u8ff0": str(product.get("tags") or "").strip(),
                "\u56fe\u7247": image,
                "\u539f\u4ef7": original_price,
                "\u6298\u6263\u4ef7": discount_price,
                "\u53d8\u4f53\u540d": "",
                "\u53d8\u4f53\u503c": variant_str,
                "\u5206\u7c7b": full_label,
            }
            row_json = json.dumps(row, ensure_ascii=False)

            if redis_conn:
                try:
                    redis_conn.rpush(redis_key_list, row_json)
                    redis_conn.incr(redis_key_count)
                except Exception:
                    pass

            with buffer_lock:
                buffer.append(row)
                if len(buffer) >= FLUSH_INTERVAL:
                    _flush_buffer_to_excel(buffer, excel_path, progress_callback)

        time.sleep(random.uniform(*PAGE_SLEEP))
        page += 1

    if progress_callback:
        progress_callback(f"  [\u5b8c\u6210] {domain} \u672c\u9875\u5339\u914d {page_all_matched} \u6761")


def _flush_buffer_to_excel(buffer, excel_path, progress_callback=None):
    if not buffer:
        return
    rows = list(buffer)
    buffer.clear()
    try:
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            for row in rows:
                ws.append([row.get(col, "") for col in EXPORT_COLUMNS])
            wb.save(excel_path)
        else:
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
            df.to_excel(excel_path, index=False)
        if progress_callback:
            progress_callback(f"  [\u5199\u5165] Excel \u5df2\u5199\u5165 {len(rows)} \u6761 (\u7d2f\u8ba1)")
    except Exception as exc:
        buffer.extend(rows)
        if progress_callback:
            progress_callback(f"  [\u9519\u8bef] Excel\u5199\u5165\u5931\u8d25: {exc}")


def _count_redis(redis_conn, key_count):
    if not redis_conn:
        return 0
    try:
        v = redis_conn.get(key_count)
        return int(v) if v else 0
    except Exception:
        return 0


def run_structured_crawl(
    domain=None,
    category_text=None,
    api_mode="5",
    progress_callback=None,
    stop_callback=None,
):
    lines = [line.strip() for line in (category_text or "").split("\\n") if line.strip()]
    categories = []
    for line in lines:
        if "|||" not in line:
            if progress_callback:
                progress_callback(f"\u8df3\u8fc7\u65e0\u6548\u5206\u7c7b\u884c: {line}")
            continue
        p, s = [x.strip() for x in line.split("|||", 1)]
        if not p or not s:
            continue
        categories.append((p, s, line))

    if not categories:
        raise ValueError("\u6ca1\u6709\u6709\u6548\u7684\u5206\u7c7b\u884c\uff0c\u683c\u5f0f: \u4e00\u7ea7\u5206\u7c7b|||\u4e8c\u7ea7\u5206\u7c7b\uff08\u6bcf\u884c\u4e00\u4e2a\uff09")

    safe_domain = re.sub(r'[\\\\/:*?"<>|]', "_", domain or "unknown")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(DATA_DIR, f"{safe_domain}_{timestamp}.xlsx")
    os.makedirs(DATA_DIR, exist_ok=True)

    redis_conn = _get_redis()
    task_id = f"{safe_domain}_{timestamp}"
    redis_key_list = _redis_key(task_id, "products")
    redis_key_count = _redis_key(task_id, "count")

    if redis_conn:
        try:
            redis_conn.delete(redis_key_list, redis_key_count)
        except Exception:
            pass

    currency_map = _load_currency_map()
    seen_unique_keys = set()
    buffer = []
    buffer_lock = threading.Lock()

    total_matched = 0
    cat_counts = {full_label: 0 for _, _, full_label in categories}
    cat_done = set()

    def get_total():
        rc = _count_redis(redis_conn, redis_key_count)
        return max(total_matched, rc)

    def update_total(delta=0):
        nonlocal total_matched
        total_matched = get_total()

    def flush():
        with buffer_lock:
            _flush_buffer_to_excel(buffer, excel_path, progress_callback)

    try:
        for idx, (primary_cat, secondary_cat, full_label) in enumerate(categories):
            if progress_callback:
                progress_callback(f"\\n===== \u5904\u7406\u5206\u7c7b ({idx+1}/{len(categories)}): {full_label} =====")

            if progress_callback:
                progress_callback(f"[\u9636\u6bb51] \u4e8c\u7ea7\u5206\u7c7b\u5339\u914d: {secondary_cat}")

            sites = _search_shopify_sites(
                keyword=secondary_cat,
                api_mode=api_mode,
                progress_callback=progress_callback,
                stop_callback=stop_callback,
            )

            if progress_callback:
                progress_callback(f"[\u9636\u6bb51] {full_label} \u53d1\u73b0 {len(sites)} \u4e2aShopify\u7ad9\u70b9")

            for site_url, site_count in sites:
                if stop_callback and stop_callback():
                    break

                current_total = get_total()
                if cat_counts[full_label] >= PER_CATEGORY_MIN and current_total >= MAX_TOTAL:
                    break

                _crawl_and_match(
                    url=site_url,
                    secondary_cat=secondary_cat,
                    primary_cat=primary_cat,
                    full_label=full_label,
                    currency_map=currency_map,
                    seen_unique_keys=seen_unique_keys,
                    redis_conn=redis_conn,
                    redis_key_list=redis_key_list,
                    redis_key_count=redis_key_count,
                    buffer=buffer,
                    buffer_lock=buffer_lock,
                    excel_path=excel_path,
                    progress_callback=progress_callback,
                    stop_callback=stop_callback,
                )
                cat_counts[full_label] = _count_redis(redis_conn, redis_key_count)
                current_total = cat_counts[full_label]

                if current_total >= PER_CATEGORY_MIN and full_label not in cat_done:
                    cat_done.add(full_label)
                    if progress_callback:
                        progress_callback(f"[\u8fbe\u6807] {full_label} \u5df2\u8fbe {current_total} \u6761")

            if stop_callback and stop_callback():
                break

            current_total = get_total()
            cat_done_count = len(cat_done)

            if cat_counts[full_label] < PER_CATEGORY_MIN:
                if progress_callback:
                    progress_callback(
                        f"[\u9636\u6bb52] {full_label} \u4ec5 {cat_counts[full_label]} \u6761\uff0c"
                        f"\u542f\u7528\u4e00\u7ea7\u5206\u7c7b\u5157\u5e95(\u9608\u503c{MATCH_THRESHOLD_PHASE2}%)"
                    )

                sites_phase2 = _search_shopify_sites(
                    keyword=primary_cat,
                    api_mode=api_mode,
                    max_results=30,
                    progress_callback=progress_callback,
                    stop_callback=stop_callback,
                )

                for site_url, site_count in sites_phase2:
                    if stop_callback and stop_callback():
                        break

                    current_total = get_total()
                    if current_total >= PER_CATEGORY_MIN and current_total >= MAX_TOTAL:
                        break

                    _crawl_and_match(
                        url=site_url,
                        secondary_cat=primary_cat,
                        primary_cat=primary_cat,
                        full_label=primary_cat,
                        currency_map=currency_map,
                        seen_unique_keys=seen_unique_keys,
                        redis_conn=redis_conn,
                        redis_key_list=redis_key_list,
                        redis_key_count=redis_key_count,
                        buffer=buffer,
                        buffer_lock=buffer_lock,
                        excel_path=excel_path,
                        progress_callback=progress_callback,
                        stop_callback=stop_callback,
                    )
                    cat_counts[full_label] = _count_redis(redis_conn, redis_key_count)
                    current_total = cat_counts[full_label]
                    if current_total >= PER_CATEGORY_MIN:
                        break

                if cat_counts[full_label] < PER_CATEGORY_MIN:
                    if progress_callback:
                        progress_callback(f"[\u4e0d\u8db3] {full_label} \u5157\u5e95\u540e\u4ecd\u4e0d\u8db3 {PER_CATEGORY_MIN} \u6761")

        all_done = all(c >= PER_CATEGORY_MIN for c in cat_counts.values())
        if all_done:
            if progress_callback:
                progress_callback("\\n\u6240\u6709\u5206\u7c7b\u5df2\u8fbe\u6807\uff0c\u7ee7\u7eed\u722c\u53d6\u81f3\u4e0a\u9650")

            for primary_cat, secondary_cat, full_label in categories:
                if stop_callback and stop_callback():
                    break

                sites_extra = _search_shopify_sites(
                    keyword=secondary_cat,
                    api_mode=api_mode,
                    progress_callback=progress_callback,
                    stop_callback=stop_callback,
                )

                for site_url, site_count in sites_extra:
                    if stop_callback and stop_callback():
                        break

                    current_total = get_total()
                    if current_total >= MAX_TOTAL:
                        break

                    _crawl_and_match(
                        url=site_url,
                        secondary_cat=secondary_cat,
                        primary_cat=primary_cat,
                        full_label=full_label,
                        currency_map=currency_map,
                        seen_unique_keys=seen_unique_keys,
                        redis_conn=redis_conn,
                        redis_key_list=redis_key_list,
                        redis_key_count=redis_key_count,
                        buffer=buffer,
                        buffer_lock=buffer_lock,
                        excel_path=excel_path,
                        progress_callback=progress_callback,
                        stop_callback=stop_callback,
                    )

                    if get_total() >= MAX_TOTAL:
                        break

        flush()
        final_total = get_total()

        if progress_callback:
            progress_callback("\\n===== \u5b8c\u6210 =====")
            progress_callback(f"\u603b\u8ba1\u5339\u914d: {final_total} \u6761")
            for fl in cat_counts:
                actual = _count_redis(redis_conn, redis_key_count)
                if actual < PER_CATEGORY_MIN:
                    progress_callback(f"  \u26a0 {fl}: \u4e0d\u8db3{PER_CATEGORY_MIN}\u6761")
            progress_callback(f"\u5bfc\u51fa\u6587\u4ef6: {excel_path}")

        if redis_conn:
            try:
                redis_conn.delete(redis_key_list, redis_key_count)
            except Exception:
                pass

        return {
            "total": final_total,
            "file_path": excel_path,
            "stopped": False,
        }

    except Exception as exc:
        flush()
        if progress_callback:
            progress_callback(f"[\u9519\u8bef] {exc}")
        return {"total": get_total(), "file_path": excel_path, "stopped": False, "error": str(exc)}
"""

target = r"C:\Users\admin\Downloads\YSQD 1 (1)\YSQD\structured_site_crawler.py"
with open(target, "w", encoding="utf-8") as f:
    f.write(content)
print("Written OK, size:", len(content), "chars")
'''

with open(r"C:\Users\admin\Downloads\YSQD 1 (1)\YSQD\_fix_temp.py", "w", encoding="utf-8") as f:
    f.write(script)

print("Temp script written")
