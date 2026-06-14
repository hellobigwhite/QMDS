import difflib
import os
import random
import re
import time
from collections import Counter, defaultdict
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from concurrent.futures import ThreadPoolExecutor, as_completed
from pymongo import ASCENDING, MongoClient, UpdateOne

try:
    import ahocorasick
    _AC_AVAILABLE = True
except ImportError:
    _AC_AVAILABLE = False


MONGO_URI = "mongodb://localhost:27017/"
SOURCE_DB = "shopify_data_new"
STAGING_DB = "shopify_data_new"
CLEAN_DB = "shopify_data_new"
RECYCLE_DB = "recycle"
RECYCLE_COLLECTION = "products_deleted"
COLLECTION_CACHE_TTL = 30

TITLE_FIELD = "标题"
DESC_FIELD = "描述"
CATEGORY_FIELD = "分类"
EXPORT_DIR = "Data/商品导出"
IMAGE_FIELD = "图片"
PRICE_FIELDS = ("折扣价", "原价", "价格")

IMAGE_REMOVE_KEYWORDS = ["coming-soon", "noimage", "default", ".svg"]
MIN_TITLE_LEN = 15
MIN_DESC_LEN = 30
MIN_PRICE = 3
MAX_PRICE = 6000
FORBIDDEN_PROGRESS_EVERY = 50000
FORBIDDEN_BATCH_SIZE = 5000
DELETE_BATCH_SIZE = 5000
WRITE_BATCH_SIZE = 2000
CATEGORY_SMALL_THRESHOLD = 30
BULK_SCAN_BATCH = 5000
PUBLIC_CATEGORY_CHOICES = [
    "Best Seller",
    "Featured",
    "Accessories",
    "Other",
    "New Arrival",
    "Exclusive",
    "Limited Edition",
    "Hot Sale",
    "Most Popular",
    "Trending",
    "Special Offer",
    "Flash Sale",
]


class CleanupStopRequested(Exception):
    pass


def _raise_if_stop_requested(stop_callback=None):
    if stop_callback and stop_callback():
        raise CleanupStopRequested("Stop requested")

PROHIBITED_KEYWORDS = [
    "gun", "firearm", "rifle", "pistol", "shotgun", "ammunition", "bullet", "silencer", "suppressor",
    "switchblade", "automatic knife", "butterfly knife", "brass knuckles", "knuckle duster", "taser", "stun gun",
    "explosive", "bomb", "grenade", "detonator", "molotov", "c4", "dynamite", "rocket", "missile",


    "drug", "cocaine", "heroin", "methamphetamine", "fentanyl", "marijuana", "cannabis", "thc",
    "cbd unapproved", "delta-8", "delta-9", "psilocybin", "lsd", "ecstasy", "mdma", "ketamine", "steroid", "anabolic",
    "opioid", "oxycodone", "vicodin", "prescription without", "rx required", "pharma restricted", "clenbuterol",
    "ephedrine", "pseudoephedrine", "dmt", "salvia", "kratom", "amanita",


    "counterfeit", "fake", "replica", "knockoff", "dupe",  "look alike", "copy", "bootleg", "pirate",
    "unauthorized", "trademark violation", "ip complaint", "high imitation",

    "hazardous", "poison", "toxic", "corrosive", "flammable", "lithium battery loose", "lithium ion loose", "mercury",
    "asbestos", "pesticide", "herbicide", "chemical weapon", "acid", "battery restricted",

    "ivory", "rhino horn", "shark fin", "turtle shell", "endangered", "cites", "wildlife product", "real fur",
    "exotic leather restricted", "python skin", "alligator",


    "cure cancer", "cure covid", "miracle cure", "fda unapproved", "medical device restricted", "cpap without",
    "prescription required",

    "lottery", "gambling", "casino", "poker chip", "slot machine", "lockpick", "lock pick", "stolen",
    "illegal activity",
    "used underwear", "human remains", "human hair unprocessed", "event ticket restricted", "embargoed goods",
    "sanctioned country", "counterfeit money", "forged", "swastika", "nazi", "forced labour",
    "child-like sex doll",
]

BROAD_PROHIBITED_KEYWORDS = {
    "alligator",
    "stolen",
    "pirate",
    "illegal activity",
    "endangered",
    "forced labour",
}

SAFE_PROHIBITED_CONTEXTS = {
    "fake": ["fake nail", "fake nails", "fake flower", "fake flowers", "fake plant", "fake plants", "fake eyelash", "fake eyelashes"],
    "copy": ["copy paper", "carbon copy", "copy stand"],
    "acid": ["acid free", "hyaluronic acid", "amino acid", "folic acid", "acid brush", "acid dye"],
    "alligator": ["alligator clip", "alligator clips"],
    "pirate": ["pirate costume", "pirate ship", "pirate hat", "pirate flag"],
}

BRAND_KEYWORDS = [
    'Hermes', 'Chanel', 'Givenchy', 'Prada', 'Gucci', 'LV', 'YSL', 'Delvaux', 'Marni', 'Mulberry', 'Dior',
    'Chloe', 'Loewe', 'Fendi', 'Proenza', 'McQueen', 'Vetements', 'Balenciaga', 'Moschino', 'Issey Miyake',
    'Canada Goose', 'Celine', 'Kenzo', 'COMME DES GAR?ONS', 'Supreme', 'Phillip Lim', 'Y-3', 'Thom Browne',
    'Coach', 'Michael Kors', 'Kate Spade', 'Under Armour', 'Tory Burch', 'Marc Jacobs', 'Armani Exchange',
    'Nike', 'Adidas', 'Louis Vuitton', 'Patek Philippe', 'Audemars Piguet', 'Vacheron Constantin',
    'Vacherron Constantin', 'A. Lange&Sohne', 'Breguet', 'Roger Dubuis', 'Parmigiani', 'Blancpain',
    'Ulysse Nardin', 'Franck Muller', 'Glashutte Original', 'Girard Perregaux', 'Rolex', 'IWC',
    'Jaeger LeCoultre', 'Cartier', 'Chopard', 'Piaget', 'Omega', 'Corum', 'Zenith', 'Movado',
    'Longines', 'Tissot', 'Seiko', 'Citizen', 'Casio', 'Bulova', 'Swatch', 'Lego',
    'Daniel Wellington', 'nintendo', 'Disney', 'Hello Kitty', 'Funko',
    'Cannabis', 'Marijuana', 'Weed', 'THC', 'CBD', 'Hash', 'Indica', 'Sativa',
    'Strain', 'Hydroponic', 'Roach Clip', 'Cocaine', 'Heroin', 'Meth', 'Methamphetamine',
    'LSD', 'Ecstasy', 'MDMA', 'Ketamine', 'Psilocybin', 'Medical', 'Drug', 'Drugs', 'Narcotic',
    'Opioid', 'Morphine', 'Fentanyl', 'Oxycodone', 'Intimate', 'Sanitary Napkin',
    'Codeine', 'Xanax', 'Diazepam', 'Valium', 'Adderall', 'Ritalin', 'Steroid',
    'Anabolic Steroid', 'Prescription Drug', 'Porn', 'Pornography', 'XXX', 'Sex', 'Sexual', 'Erotic',
    'Erotica', 'Fetish', 'BDSM', 'Escort', 'Prostitution', 'Stripper', 'Strip Club',
    'Nude', 'Nudity', 'OnlyFans', 'Camgirl', 'Cam Boy', 'Gun', 'Guns', 'Firearm', 'Pistol', 'Rifle',
    'Shotgun', 'Revolver', 'Ammunition', 'Ammo', 'Bullet', 'Magazine', 'Silencer', 'Suppressor',
    'Dagger', 'Sword', 'Machete', 'Crossbow', 'Explosive', 'Bomb', 'Grenade', 'Weapon', 'Weapons', 'Hemp',
    'Hermes', 'Chanel', 'Givenchy', 'Prada', 'Bvlgari', 'Miu Miu',
    'Christian Dior', 'Saint Laurent', 'Bottega Veneta', 'Valentino', 'Alexander McQueen',
    'Moncler', 'Lacoste', 'Ralph Lauren', 'Hugo Boss', 'Off White', 'The Row', 'Acne Studios',
    'Jil Sander', 'Dries Van Noten', 'Furla', 'Rimowa', 'Goyard', 'Sophie Hulme', 'Tods', 'Brunello Cucinelli',
    'Philipp Plein', 'Balmain', 'Stella McCartney', 'Isabel Marant', 'Rei Kawakubo',
]
_BRAND_PATTERNS = [r'\b' + re.escape(kw) + r'\b' for kw in dict.fromkeys(BRAND_KEYWORDS)]
_BRAND_REGEX = re.compile('|'.join(_BRAND_PATTERNS), re.IGNORECASE)

APPAREL_CATEGORY_KEYWORDS = [
    "apparel",
    "clothing",
    "garment",
    "outerwear",
    "activewear",
    "sportswear",
    "streetwear",
    "underwear",
    "lingerie",
    "sleepwear",
    "dress",
    "dresses",
    "skirt",
    "skirts",
    "shirt",
    "shirts",
    "t shirt",
    "t-shirt",
    "tee",
    "blouse",
    "hoodie",
    "hoodies",
    "sweatshirt",
    "sweatshirts",
    "sweater",
    "sweaters",
    "jacket",
    "jackets",
    "coat",
    "coats",
    "pants",
    "trousers",
    "jeans",
    "leggings",
    "shorts",
    "jumpsuit",
    "jumpsuits",
    "romper",
    "rompers",
    "shoe",
    "shoes",
    "sneaker",
    "sneakers",
    "boot",
    "boots",
    "sandal",
    "sandals",
    "slipper",
    "slippers",
    "heel",
    "heels",
]

BRAND_TO_CATEGORY: Dict[str, str] = {
    "louis vuitton": "Luxury Handbags",
    "gucci": "Luxury Handbags & Accessories",
    "hermes": "Luxury Leather Goods",
    "prada": "Luxury Handbags",
    "chanel": "Luxury Fashion",
    "dior": "Luxury Fashion & Beauty",
    "balenciaga": "Luxury Streetwear",
    "saint laurent": "Luxury Fashion",
    "ysl": "Luxury Fashion",
    "burberry": "Luxury Trench & Fashion",
    "celine": "Luxury Minimalist Bags",
    "fendi": "Luxury Handbags",
    "loewe": "Luxury Handbags",
    "bottega veneta": "Luxury Leather Goods",
    "alexander mcqueen": "Luxury Fashion",
    "miu miu": "Luxury Fashion",
    "the row": "Luxury Minimalist Fashion",
    "moncler": "Luxury Outerwear",
    "versace": "Luxury Fashion",
    "rolex": "Luxury Watches",
    "patek philippe": "Luxury Watches",
    "audemars piguet": "Luxury Watches",
    "vacheron constantin": "Luxury Watches",
    "richard mille": "Luxury Watches",
    "jaeger-lecoultre": "Luxury Watches",
    "breitling": "Luxury Watches",
    "omega": "Luxury Watches",
    "cartier": "Luxury Jewelry & Watches",
    "tiffany": "Luxury Jewelry",
    "coach": "Mid-Range Handbags",
    "kate spade": "Mid-Range Handbags",
    "tory burch": "Mid-Range Handbags",
    "michael kors": "Mid-Range Handbags",
    "furla": "Mid-Range Handbags",
    "mcm": "Mid-Range Handbags",
    "polene": "Mid-Range Handbags",
    "demellier": "Mid-Range Handbags",
    "strathberry": "Mid-Range Handbags",
    "staud": "Mid-Range Handbags",
    "by far": "Mid-Range Handbags",
    "osoi": "Mid-Range Handbags",
    "savette": "Mid-Range Handbags",
    "khaite": "Mid-Range Handbags",
    "ganni": "Contemporary Fashion",
    "nanushka": "Contemporary Fashion",
    "sezane": "Contemporary Fashion",
    "toteme": "Contemporary Fashion",
    "sandro": "Contemporary Fashion",
    "maje": "Contemporary Fashion",
    "reformation": "Contemporary Fashion",
    "everlane": "Contemporary Fashion",
    "melissa & doug": "Mid-Range Educational Toys",
    "fat brain toys": "Mid-Range Educational Toys",
    "little tikes": "Mid-Range Outdoor Toys",
    "vtech": "Mid-Range Learning Toys",
    "playskool": "Mid-Range Starter Toys",
    "jellycat": "Mid-Range Plush Toys",
    "pop mart": "Mid-Range Collectible Blind Boxes",
    "bubble mart": "Mid-Range Collectible Blind Boxes",
    "52toys": "Mid-Range Collectible Blind Boxes",
    "sonny angel": "Mid-Range Collectible Figures",
    "labubu": "Mid-Range Collectible Blind Boxes",
    "warmies": "Mid-Range Weighted Plush",
    "hape": "Mid-Range Wooden Toys",
    "mideer": "Mid-Range Puzzle Toys",
}

_COLLECTION_CACHE = {"time": 0.0, "items": []}
_ENGLISH_PROHIBITED_PATTERNS = []
_CHINESE_PROHIBITED_KEYWORDS = []
_CHINESE_PROHIBITED_REGEX = None
_MONGO_ENGLISH_PROHIBITED = []
_ACCENT_REGEX = None
_APPAREL_CATEGORY_PATTERNS = []
_BRAND_PATTERNS = []
_SAFE_PROHIBITED_CONTEXTS_NORMALIZED = {
    keyword: tuple(phrase.lower().replace("-", " ") for phrase in phrases)
    for keyword, phrases in SAFE_PROHIBITED_CONTEXTS.items()
}
_NON_LATIN_RE = re.compile(
    "[\u0400-\u04FF"   # Cyrillic
    "\u0590-\u05FF"    # Hebrew
    "\u0600-\u06FF"    # Arabic
    "\u0900-\u097F"    # Devanagari
    "\u0E00-\u0E7F"    # Thai
    "\u3040-\u30FF"    # Hiragana + Katakana
    "\u3400-\u4DBF"    # CJK Extension A
    "\u4E00-\u9FFF"    # CJK
    "\uAC00-\uD7AF"    # Hangul
    "]"
)
_LETTER_RE = re.compile(r'[^\W\d_]', re.UNICODE)
_LATIN_RE = re.compile(r'[A-Za-z]')

_NORMALIZED_PROHIBITED_KEYWORDS = tuple(
    dict.fromkeys(
        keyword.lower().strip().replace("-", " ")
        for keyword in PROHIBITED_KEYWORDS
        if keyword and keyword.strip()
    )
)

for lowered in _NORMALIZED_PROHIBITED_KEYWORDS:
    if not lowered:
        continue
    if re.search(r"[\u4e00-\u9fff]", lowered):
        _CHINESE_PROHIBITED_KEYWORDS.append(lowered)
    else:
        _ENGLISH_PROHIBITED_PATTERNS.append(
            (lowered, re.compile(r"(?<![a-z0-9])" + re.escape(lowered) + r"(?![a-z0-9])", re.IGNORECASE))
        )

if _CHINESE_PROHIBITED_KEYWORDS:
    _CHINESE_PROHIBITED_REGEX = re.compile(
        "|".join(re.escape(kw) for kw in sorted(_CHINESE_PROHIBITED_KEYWORDS, key=len, reverse=True))
    )

_MONGO_ENGLISH_PROHIBITED = [
    (kw, "(^|[^a-z0-9])" + re.escape(kw) + "([^a-z0-9]|$)")
    for kw, _ in _ENGLISH_PROHIBITED_PATTERNS
]
_ENGLISH_PROHIBITED_REGEX = re.compile(
    "|".join("(?:" + pat + ")" for _, pat in _MONGO_ENGLISH_PROHIBITED),
    re.IGNORECASE
) if _MONGO_ENGLISH_PROHIBITED else None

_ENGLISH_COMBINED_RE = re.compile(
    "(?<![a-z0-9])(" + "|".join(
        re.escape(kw) for kw, _ in sorted(_ENGLISH_PROHIBITED_PATTERNS, key=lambda x: -len(x[0]))
    ) + ")(?![a-z0-9])",
    re.IGNORECASE
) if _ENGLISH_PROHIBITED_PATTERNS else None

_AC_AUTOMATON = None
if _AC_AVAILABLE and (_CHINESE_PROHIBITED_KEYWORDS or _ENGLISH_PROHIBITED_PATTERNS):
    _ac = ahocorasick.Automaton()
    for kw in _CHINESE_PROHIBITED_KEYWORDS:
        _ac.add_word(kw, ('cn', kw))
    for kw, _ in _ENGLISH_PROHIBITED_PATTERNS:
        _ac.add_word(kw, ('en', kw))
    _ac.make_automaton()
    _AC_AUTOMATON = _ac
    del _ac

for keyword in dict.fromkeys(
    keyword.lower().strip().replace("-", " ")
    for keyword in APPAREL_CATEGORY_KEYWORDS
    if keyword and keyword.strip()
):
    if re.search(r"[\u4e00-\u9fff]", keyword):
        pattern = re.compile(re.escape(keyword), re.IGNORECASE)
    else:
        pattern = re.compile(r"(?<![a-z0-9])" + re.escape(keyword) + r"(?![a-z0-9])", re.IGNORECASE)
    _APPAREL_CATEGORY_PATTERNS.append((keyword, pattern))


def get_client():
    return MongoClient(MONGO_URI)


def get_source_db(client):
    return client[SOURCE_DB]


def ensure_source_indexes(db):
    for name in db.list_collection_names():
        coll = db[name]
        existing = coll.index_information()
        existing_key_sets = {tuple(v["key"]) for v in existing.values()}
        for field in ("unique_key", "source_url", "crawl_time"):
            target = tuple([(field, 1)])
            if target not in existing_key_sets:
                coll.create_index(field, name=f"idx_{field}")


def ensure_collection_indexes(collection, fields):
    existing = collection.index_information()
    existing_key_sets = {tuple(v["key"]) for v in existing.values()}
    for field in fields:
        target = tuple([(field, ASCENDING)])
        if target not in existing_key_sets:
            safe_name = re.sub(r"[^A-Za-z0-9_]+", "_", field).strip("_") or "field"
            collection.create_index(field, name=f"idx_{safe_name}")


def invalidate_collection_cache():
    _COLLECTION_CACHE["time"] = 0.0
    _COLLECTION_CACHE["items"] = []


def list_product_collections():
    now = time.time()
    if _COLLECTION_CACHE["items"] and now - _COLLECTION_CACHE["time"] < COLLECTION_CACHE_TTL:
        return list(_COLLECTION_CACHE["items"])

    client = get_client()
    try:
        db = get_source_db(client)
        items = []
        for name in sorted(db.list_collection_names()):
            try:
                count = db[name].estimated_document_count()
            except Exception:
                count = db[name].count_documents({})
            items.append({"name": name, "count": count})
        _COLLECTION_CACHE["time"] = now
        _COLLECTION_CACHE["items"] = list(items)
        return items
    finally:
        client.close()


def resolve_collections(db, selected_collection):
    selected = (selected_collection or "").strip()
    names = sorted(db.list_collection_names())
    if selected == "__all__":
        return names
    if not selected:
        return []
    return [selected] if selected in names else []


def describe_collection_scope(selected_collection, collections, skip_clean=False):
    selected = (selected_collection or "").strip()
    effective = [name for name in collections if not (skip_clean and name.endswith("_clean"))]
    if selected == "__all__":
        return f"Scope: all collections; effective collection count: {len(effective)}"
    if not selected:
        return "Scope is empty; refused to fall back to all collections"
    if not collections:
        return f"Scope: {selected}; collection not found, no data will be processed"
    return f"Scope: {selected}; effective collections: {', '.join(effective) if effective else 'none'}"


def delete_ids_in_batches(collection, ids, batch_size=DELETE_BATCH_SIZE):
    deleted = 0
    for start in range(0, len(ids), batch_size):
        chunk = ids[start : start + batch_size]
        if not chunk:
            continue
        deleted += collection.delete_many({"_id": {"$in": chunk}}).deleted_count
    return deleted


def bulk_write_in_batches(collection, operations, batch_size=WRITE_BATCH_SIZE):
    affected = 0
    for start in range(0, len(operations), batch_size):
        chunk = operations[start : start + batch_size]
        if not chunk:
            continue
        result = collection.bulk_write(chunk, ordered=False)
        affected += int(getattr(result, "modified_count", 0))
    return affected


def insert_docs_in_batches(collection, docs, batch_size=WRITE_BATCH_SIZE):
    inserted = 0
    for start in range(0, len(docs), batch_size):
        chunk = docs[start : start + batch_size]
        if not chunk:
            continue
        collection.insert_many(chunk, ordered=False)
        inserted += len(chunk)
    return inserted


def write_clean_docs_batch(collection, docs, batch_size=WRITE_BATCH_SIZE):
    inserted = 0
    existing = 0
    keyed_ops = []
    unkeyed_docs = []
    seen_keys = set()

    for doc in docs:
        unique_key = text_value(doc.get("unique_key"))
        if not unique_key:
            unkeyed_docs.append(doc)
            continue
        if unique_key in seen_keys:
            existing += 1
            continue
        seen_keys.add(unique_key)
        keyed_ops.append(UpdateOne({"unique_key": unique_key}, {"$setOnInsert": doc}, upsert=True))

    for start in range(0, len(keyed_ops), batch_size):
        chunk = keyed_ops[start : start + batch_size]
        if not chunk:
            continue
        result = collection.bulk_write(chunk, ordered=False)
        inserted += int(getattr(result, "upserted_count", 0))
        existing += int(getattr(result, "matched_count", 0))

    if unkeyed_docs:
        inserted += insert_docs_in_batches(collection, unkeyed_docs, batch_size=batch_size)

    return inserted, existing


def text_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_price(doc):
    for field in PRICE_FIELDS:
        value = doc.get(field)
        if value in ("", None):
            continue
        try:
            return float(value)
        except Exception:
            continue
    return None


def get_basic_delete_reason(doc):
    title = text_value(doc.get(TITLE_FIELD))
    desc = text_value(doc.get(DESC_FIELD))
    if not title or not desc:
        return "empty"
    if len(desc) < MIN_DESC_LEN:
        return "short_desc"
    if len(title) < MIN_TITLE_LEN:
        return "short_title"
    if re.fullmatch(r"\s*\d+\s*", title):
        return "numeric_title"
    price = parse_price(doc)
    if price is not None and (price < MIN_PRICE or price > MAX_PRICE):
        return "bad_price"
    return ""


def has_bad_image(doc):
    image = doc.get(IMAGE_FIELD)
    if image in (None, "", []):
        return True
    text = str(image).strip().lower()
    if not text:
        return True
    return any(keyword in text for keyword in IMAGE_REMOVE_KEYWORDS)


def has_brand_keyword(doc) -> bool:
    text = " ".join(
        text_value(doc.get(field)) for field in (TITLE_FIELD, CATEGORY_FIELD, DESC_FIELD)
        if text_value(doc.get(field))
    )
    if not text:
        return False
    return bool(_BRAND_REGEX.search(text))


def strip_html(text: str) -> str:
    value = text_value(text)
    if not value:
        return ""
    value = re.sub(r"<[^>]+>", " ", value)
    value = re.sub(r"&[a-z0-9#]+;", " ", value, flags=re.IGNORECASE)
    return " ".join(value.split()).strip()


def has_non_latin_script(text: str) -> bool:
    value = text_value(text)
    if not value:
        return False
    return bool(_NON_LATIN_RE.search(value))


NON_ENGLISH_EUROPEAN_CHARS = set(
    "àáâãäåæçèéêëìíîïðñòóôõöøùúûüýþÿ"
    "ÀÁÂÃÄÅÆÇÈÉÊËÌÍÎÏÐÑÒÓÔÕÖØÙÚÛÜÝÞß"
)

_ACCENT_REGEX = "[" + re.escape("".join(sorted(NON_ENGLISH_EUROPEAN_CHARS))) + "]"


def has_european_non_english(text: str) -> bool:
    if not text:
        return False
    return any(ch in NON_ENGLISH_EUROPEAN_CHARS for ch in text)


def get_non_english_reason(doc) -> str:
    title = strip_html(doc.get(TITLE_FIELD))
    category = strip_html(doc.get(CATEGORY_FIELD))
    desc = strip_html(doc.get(DESC_FIELD))
    primary_text = " ".join(part for part in [title, category] if part).strip()
    combined = " ".join(part for part in [title, category, desc] if part).strip()
    if not primary_text and not combined:
        return ""

    if has_non_latin_script(primary_text):
        return "non_latin_primary"

    if has_non_latin_script(desc):
        primary_letters = re.findall(r"[^\W\d_]", primary_text, re.UNICODE)
        primary_latin_letters = re.findall(r"[A-Za-z]", primary_text)
        if len(primary_letters) < 8:
            return "non_latin_desc_with_weak_primary"
        if primary_letters:
            primary_latin_ratio = len(primary_latin_letters) / max(len(primary_letters), 1)
            if primary_latin_ratio < 0.5:
                return "non_latin_desc_with_weak_primary"

    primary_letters = re.findall(r"[^\W\d_]", primary_text, re.UNICODE)
    primary_latin_letters = re.findall(r"[A-Za-z]", primary_text)
    if len(primary_letters) >= 12:
        primary_latin_ratio = len(primary_latin_letters) / max(len(primary_letters), 1)
        if primary_latin_ratio < 0.55:
            return "low_latin_primary"

    combined_letters = re.findall(r"[^\W\d_]", combined, re.UNICODE)
    combined_latin_letters = re.findall(r"[A-Za-z]", combined)
    if len(combined_letters) >= 80:
        combined_latin_ratio = len(combined_latin_letters) / max(len(combined_letters), 1)
        if combined_latin_ratio < 0.45 and len(primary_latin_letters) < 8:
            return "low_latin_combined"

    if has_european_non_english(primary_text):
        return "european_accent"

    return ""


def is_mostly_chinese_category(text: str) -> bool:
    value = text_value(text)
    if not value:
        return False

    cleaned = re.sub(r"[\s\|\>\-:/\\,&;]+", "", value)
    if not cleaned:
        return False

    chinese_count = len(re.findall(r"[\u4e00-\u9fff]", cleaned))
    english_count = len(re.findall(r"[A-Za-z]", cleaned))
    digit_count = len(re.findall(r"\d", cleaned))
    valid_char_count = chinese_count + english_count + digit_count
    if valid_char_count == 0:
        return False

    chinese_ratio = chinese_count / valid_char_count
    return chinese_count > 0 and chinese_ratio >= 0.8 and english_count == 0


def is_safe_prohibited_context(keyword: str, text: str) -> bool:
    lowered = text_value(text).lower().replace("-", " ")
    if not lowered:
        return False
    for phrase in _SAFE_PROHIBITED_CONTEXTS_NORMALIZED.get(keyword, ()):
        if phrase in lowered:
            return True
    return False


def find_prohibited_keywords(text: str, first_only: bool = False) -> List[str]:
    value = text_value(text)
    if not value:
        return []

    lowered = value.lower().replace("-", " ")
    matched = []

    if _AC_AUTOMATON:
        seen = set()
        for end_idx, (lang, kw) in _AC_AUTOMATON.iter(lowered):
            if lang == 'en':
                start = end_idx + 1 - len(kw)
                prev_ok = start == 0 or not lowered[start - 1].isalnum()
                next_ok = end_idx + 1 >= len(lowered) or not lowered[end_idx + 1].isalnum()
                if not (prev_ok and next_ok):
                    continue
            if kw in seen:
                continue
            seen.add(kw)
            if not is_safe_prohibited_context(kw, lowered):
                matched.append(kw)
                if first_only:
                    return matched
    else:
        if _CHINESE_PROHIBITED_REGEX:
            seen = set()
            for kw in _CHINESE_PROHIBITED_REGEX.findall(lowered):
                if kw in seen:
                    continue
                seen.add(kw)
                if not is_safe_prohibited_context(kw, lowered):
                    matched.append(kw)
                    if first_only:
                        return matched
        if _ENGLISH_COMBINED_RE:
            seen = set()
            for m in _ENGLISH_COMBINED_RE.finditer(lowered):
                kw = m.group(1)
                if kw in seen:
                    continue
                seen.add(kw)
                if not is_safe_prohibited_context(kw, lowered):
                    matched.append(kw)
                    if first_only:
                        return matched
    return matched


def _field_has_prohibited_quick(text) -> bool:
    """Fast pre-check: does this field text contain ANY prohibited keyword pattern?
    Only used as fallback when Aho-Corasick is unavailable."""
    value = text_value(text)
    if not value:
        return False
    lowered = value.lower().replace("-", " ")
    if _CHINESE_PROHIBITED_REGEX and _CHINESE_PROHIBITED_REGEX.search(lowered):
        return True
    if _ENGLISH_COMBINED_RE and _ENGLISH_COMBINED_RE.search(lowered):
        return True
    return False


def find_prohibited_match(doc) -> Optional[Tuple[str, str]]:
    broad_desc_matches = []

    for field in (TITLE_FIELD, CATEGORY_FIELD, DESC_FIELD):
        text = doc.get(field)
        if _AC_AUTOMATON is None and not _field_has_prohibited_quick(text):
            continue
        keywords = find_prohibited_keywords(text)
        for keyword in keywords:
            if keyword in BROAD_PROHIBITED_KEYWORDS:
                if field in {TITLE_FIELD, CATEGORY_FIELD}:
                    return field, keyword
                broad_desc_matches.append((field, keyword))
            else:
                return field, keyword

    distinct_broad_keywords = []
    seen = set()
    for field, keyword in broad_desc_matches:
        if keyword not in seen:
            seen.add(keyword)
            distinct_broad_keywords.append((field, keyword))

    if len(distinct_broad_keywords) >= 1:
        return distinct_broad_keywords[0]
    apparel_keyword = find_apparel_category_match(doc)
    if apparel_keyword:
        return CATEGORY_FIELD, apparel_keyword
    return None


def find_apparel_category_match(doc) -> str:
    texts = [doc.get(CATEGORY_FIELD), doc.get("source_category")]
    for text in texts:
        lowered = clean_for_match(text)
        if not lowered:
            continue
        for keyword, pattern in _APPAREL_CATEGORY_PATTERNS:
            if pattern.search(lowered):
                return f"apparel:{keyword}"
    return ""


def clean_for_match(text: str) -> str:
    value = text_value(text).lower()
    if not value:
        return ""
    for item in ["|||", "|", ">", "->", "→", ":", "/", "\\", "\n", "\r", "\t", "&"]:
        value = value.replace(item, " ")
    value = " ".join(value.split())
    return value.strip()


def singularize_word(word: str) -> str:
    if not word:
        return word

    value = word.strip()
    lowered = value.lower()
    irregular_map = {
        "children": "child",
        "men": "man",
        "women": "woman",
        "people": "person",
        "teeth": "tooth",
        "feet": "foot",
        "geese": "goose",
        "mice": "mouse",
    }
    if lowered in irregular_map:
        return irregular_map[lowered]
    if len(value) <= 3:
        return value
    if lowered.endswith("ies") and len(value) > 4:
        return value[:-3] + "y"
    if lowered.endswith(("ches", "shes", "xes", "zes", "sses")) and len(value) > 4:
        return value[:-2]
    if lowered.endswith("s") and not lowered.endswith(("ss", "us", "is")) and len(value) > 3:
        return value[:-1]
    return value


def normalize_category_token(token: str) -> str:
    value = text_value(token).lower()
    if not value:
        return ""
    return singularize_word(value)


def normalize_category_key(text: str) -> str:
    value = text_value(text)
    if not value:
        return ""

    parts = [part.strip() for part in value.split("|||") if part.strip()]
    normalized_parts = []
    for part in parts:
        words = re.findall(r"[A-Za-z0-9]+", part.lower())
        if words:
            normalized_part = " ".join(
                normalize_category_token(word) for word in words if text_value(word)
            ).strip()
        else:
            normalized_part = part.strip().lower()
        if normalized_part:
            normalized_parts.append(normalized_part)
    return "|||".join(normalized_parts)


def get_best_category_match(small_cat: str, large_cats: List[str]) -> str:
    if not large_cats:
        return "Other"

    small_norm = normalize_category_key(small_cat)
    if not small_norm:
        return "Other"

    for large in large_cats:
        if small_norm == normalize_category_key(large):
            return large

    best_match = None
    best_score = -1.0
    small_words = set(clean_for_match(small_norm).split())
    if not small_words:
        return "Other"

    for large in large_cats:
        large_norm = normalize_category_key(large)
        large_clean = clean_for_match(large_norm)
        if not large_clean:
            continue

        if small_norm in large_norm or large_clean.startswith(clean_for_match(small_norm) + " "):
            return large

        large_words = set(large_clean.split())
        intersection = len(small_words & large_words)
        union = len(small_words | large_words)
        if union == 0:
            continue

        jaccard = intersection / union
        len_penalty = min(len(small_words), len(large_words)) / max(len(small_words), len(large_words))
        score = jaccard * (0.75 + 0.25 * len_penalty)
        if intersection >= 1 and score > best_score:
            best_score = score
            best_match = large

    if best_match and best_score >= 0.40:
        return best_match

    cleaned_map = {}
    for category in large_cats:
        cleaned = clean_for_match(normalize_category_key(category))
        if cleaned and cleaned not in cleaned_map:
            cleaned_map[cleaned] = category

    matches = difflib.get_close_matches(clean_for_match(small_norm), list(cleaned_map.keys()), n=1, cutoff=0.62)
    if matches:
        return cleaned_map[matches[0]]

    return "Other"


def clean_category(text):
    value = text_value(text)
    if not value:
        return "Other"

    value = value.replace("_", " ")
    cleaned_symbols = "".join(char for char in value.lower() if char not in " &|>:/\\\t\n\r")
    if not cleaned_symbols or cleaned_symbols.isdigit():
        return "Other"

    for sep in ["->", "→", ">", ":", "/", "\\", "|", "&", "&&", "&amp;", "|||", "\r\n", "\n", "\r", "\t", ",", ";"]:
        value = value.replace(sep, "|||")

    value = "|||".join([part.strip() for part in value.split("|||") if part.strip()]).strip("||| ").strip()
    parts = [part.strip() for part in value.split("|||") if part.strip()]
    if not parts:
        return "Other"

    if parts and parts[0].replace(".", "").replace("-", "").isdigit():
        parts = parts[1:]

    garbage = {
        "", "-", "--", "---", "none", "null", "unknown", "other", "others",
        "na", "n/a", "test", "demo", "temp",
    }
    parts = [part for part in parts if part.lower() not in garbage and len(part) > 1]
    if len(parts) == 1 and len(parts[0]) <= 2 and not any(char.isalpha() for char in parts[0]):
        return "Other"

    meaningless_top = {
        "home", "root", "main", "category", "categories", "top", "uncategorized",
        "other", "others", "misc", "miscellaneous", "all", "everything",
    }
    if parts and parts[0].lower() in meaningless_top:
        parts = parts[1:]
    if not parts:
        return "Other"

    joined_text = "|||".join(parts)
    if is_mostly_chinese_category(joined_text):
        return "DROP_ROW"

    replaced_parts = []
    for part in parts:
        lowered = part.lower()
        replaced = None
        for pattern, category in _BRAND_PATTERNS:
            if pattern.search(lowered):
                replaced = category
                break
        replaced_parts.append(replaced or part)

    if not replaced_parts:
        return "Other"
    return "|||".join(replaced_parts)


def run_basic_cleanup(selected_collection=None, progress_callback=None, stop_callback=None):
    client = get_client()
    try:
        db = get_source_db(client)
        collections = resolve_collections(db, selected_collection)
        total_deleted = 0

        for name in collections:
            _raise_if_stop_requested(stop_callback)
            collection = db[name]
            deleted_ids = []
            reasons = Counter()

            projection = {TITLE_FIELD: 1, DESC_FIELD: 1}
            for field in PRICE_FIELDS:
                projection[field] = 1

            for doc in collection.find({}, projection):
                _raise_if_stop_requested(stop_callback)
                reason = get_basic_delete_reason(doc)
                if reason:
                    deleted_ids.append(doc["_id"])
                    reasons[reason] += 1

            deleted = delete_ids_in_batches(collection, deleted_ids) if deleted_ids else 0
            total_deleted += deleted
            if progress_callback:
                progress_callback(
                    f"[{name}] 基础数据清洗完成，删除 {deleted} 条"
                    f" | 空标题/描述 {reasons.get('empty', 0)}"
                    f" | 短标题 {reasons.get('short_title', 0)}"
                    f" | 纯数字标题 {reasons.get('numeric_title', 0)}"
                    f" | 价格异常 {reasons.get('bad_price', 0)}"
                )

        invalidate_collection_cache()
        return {"collections": len(collections), "deleted": total_deleted, "stopped": False}
    except CleanupStopRequested:
        return {"collections": len(collections), "deleted": total_deleted, "stopped": True}
    finally:
        client.close()


def run_image_cleanup(selected_collection=None, progress_callback=None, stop_callback=None):
    client = get_client()
    try:
        db = get_source_db(client)
        collections = resolve_collections(db, selected_collection)
        total_deleted = 0

        for name in collections:
            _raise_if_stop_requested(stop_callback)
            collection = db[name]
            deleted_ids = []
            for doc in collection.find({}, {IMAGE_FIELD: 1}):
                _raise_if_stop_requested(stop_callback)
                if has_bad_image(doc):
                    deleted_ids.append(doc["_id"])

            deleted = delete_ids_in_batches(collection, deleted_ids) if deleted_ids else 0
            total_deleted += deleted
            if progress_callback:
                progress_callback(f"[{name}] 异常图片清洗完成，删除 {deleted} 条")

        invalidate_collection_cache()
        return {"collections": len(collections), "deleted": total_deleted, "stopped": False}
    except CleanupStopRequested:
        return {"collections": len(collections), "deleted": total_deleted, "stopped": True}
    finally:
        client.close()


def run_english_cleanup(selected_collection=None, progress_callback=None, stop_callback=None):
    client = get_client()
    try:
        db = get_source_db(client)
        collections = resolve_collections(db, selected_collection)
        total_deleted = 0

        for name in collections:
            _raise_if_stop_requested(stop_callback)
            collection = db[name]
            deleted_ids = []
            reasons = Counter()
            projection = {TITLE_FIELD: 1, CATEGORY_FIELD: 1, DESC_FIELD: 1}

            for doc in collection.find({}, projection):
                _raise_if_stop_requested(stop_callback)
                reason = get_non_english_reason(doc)
                if reason:
                    deleted_ids.append(doc["_id"])
                    reasons[reason] += 1

            deleted = delete_ids_in_batches(collection, deleted_ids) if deleted_ids else 0
            total_deleted += deleted
            if progress_callback:
                progress_callback(
                    f"[{name}] 英文数据过滤完成，删除 {deleted} 条"
                    f" | 标题/分类含非拉丁文字 {reasons.get('non_latin_primary', 0)}"
                    f" | 描述非拉丁且主字段过弱 {reasons.get('non_latin_desc_with_weak_primary', 0)}"
                    f" | 主字段英文占比过低 {reasons.get('low_latin_primary', 0)}"
                    f" | 全文英文占比过低 {reasons.get('low_latin_combined', 0)}"
                    f" | 欧洲非英语重音字符 {reasons.get('european_accent', 0)}"
                )

        invalidate_collection_cache()
        return {"collections": len(collections), "deleted": total_deleted, "stopped": False}
    except CleanupStopRequested:
        return {"collections": len(collections), "deleted": total_deleted, "stopped": True}
    finally:
        client.close()


def run_forbidden_cleanup(selected_collection=None, progress_callback=None, stop_callback=None):
    client = get_client()
    try:
        db = get_source_db(client)
        recycle_collection = client[RECYCLE_DB][RECYCLE_COLLECTION]
        collections = resolve_collections(db, selected_collection)
        total_moved = 0

        for name in collections:
            _raise_if_stop_requested(stop_callback)
            source_collection = db[name]
            estimated = source_collection.estimated_document_count()
            if progress_callback:
                progress_callback(f"[{name}] 开始违禁词过滤，预计扫描 {estimated} 条")

            scanned = 0
            moved = 0
            matched_docs = []
            matched_ids = []

            projection = {
                CATEGORY_FIELD: 1,
                TITLE_FIELD: 1,
                DESC_FIELD: 1,
                IMAGE_FIELD: 1,
                "source_url": 1,
                "source_category": 1,
            }

            for doc in source_collection.find({}, projection):
                _raise_if_stop_requested(stop_callback)
                scanned += 1
                match = find_prohibited_match(doc)
                if match:
                    match_field, match_keyword = match
                    recycle_doc = dict(doc)
                    original_id = recycle_doc.pop("_id", None)
                    recycle_doc["original_id"] = str(original_id) if original_id is not None else ""
                    recycle_doc["recycle_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    recycle_doc["recycle_reason"] = "apparel_category" if str(match_keyword).startswith("apparel:") else "forbidden_keyword"
                    recycle_doc["recycle_source_collection"] = name
                    recycle_doc["recycle_match_field"] = match_field
                    recycle_doc["recycle_match_keyword"] = match_keyword
                    matched_docs.append(recycle_doc)
                    matched_ids.append(doc["_id"])

                if len(matched_docs) >= FORBIDDEN_BATCH_SIZE:
                    recycle_collection.insert_many(matched_docs, ordered=False)
                    moved += delete_ids_in_batches(source_collection, matched_ids)
                    matched_docs = []
                    matched_ids = []
                    if progress_callback:
                        progress_callback(f"[{name}] 违禁词过滤处理中，已扫描 {scanned} 条，已移入回收站 {moved} 条")
                elif progress_callback and scanned % FORBIDDEN_PROGRESS_EVERY == 0:
                    progress_callback(f"[{name}] 违禁词过滤处理中，已扫描 {scanned} 条，当前已命中 {moved + len(matched_ids)} 条")

            _raise_if_stop_requested(stop_callback)
            if matched_docs:
                recycle_collection.insert_many(matched_docs, ordered=False)
                moved += delete_ids_in_batches(source_collection, matched_ids)

            total_moved += moved
            if progress_callback:
                progress_callback(f"[{name}] 违禁词过滤完成，移入回收站 {moved} 条")

        invalidate_collection_cache()
        return {"collections": len(collections), "moved": total_moved, "stopped": False}
    except CleanupStopRequested:
        return {"collections": results, "stopped": True}
    finally:
        client.close()


def resolve_clean_collections(db, selected_collection):
    selected = (selected_collection or "").strip()
    clean_collections = sorted(name for name in db.list_collection_names() if name.endswith("_clean"))
    if not selected or selected == "__all__":
        return clean_collections
    if not selected.endswith("_clean"):
        return []
    return [selected] if selected in clean_collections else []


def run_clean_forbidden_cleanup(selected_collection=None, progress_callback=None, stop_callback=None):
    client = get_client()
    try:
        db = get_source_db(client)
        collections = resolve_clean_collections(db, selected_collection)
        total_deleted = 0

        if not collections and progress_callback:
            progress_callback("未找到可过滤的 _clean 集合")

        projection = {
            CATEGORY_FIELD: 1,
            TITLE_FIELD: 1,
            DESC_FIELD: 1,
            "source_url": 1,
            "source_category": 1,
        }

        for name in collections:
            _raise_if_stop_requested(stop_callback)
            collection = db[name]
            estimated = collection.estimated_document_count()
            if progress_callback:
                progress_callback(f"[{name}] 开始 clean 集合违禁词过滤，预计扫描 {estimated} 条")

            scanned = 0
            matched_ids = []
            deleted = 0

            for doc in collection.find({}, projection, batch_size=BULK_SCAN_BATCH):
                _raise_if_stop_requested(stop_callback)
                scanned += 1
                match = find_prohibited_match(doc)
                if not match:
                    if progress_callback and scanned % FORBIDDEN_PROGRESS_EVERY == 0:
                        progress_callback(f"[{name}] 已扫描 {scanned} 条，当前命中 {len(matched_ids) + deleted} 条")
                    continue

                match_field, match_keyword = match
                matched_ids.append(doc["_id"])
                if progress_callback:
                    title = text_value(doc.get(TITLE_FIELD))[:120]
                    category = text_value(doc.get(CATEGORY_FIELD))[:120]
                    source_url = text_value(doc.get("source_url"))[:200]
                    progress_callback(
                        f"[{name}] 命中并删除 | 字段={match_field} | 关键词={match_keyword} "
                        f"| 标题={title or '-'} | 分类={category or '-'} | URL={source_url or '-'}"
                    )

                if len(matched_ids) >= DELETE_BATCH_SIZE:
                    deleted += delete_ids_in_batches(collection, matched_ids)
                    matched_ids = []
                    if progress_callback:
                        progress_callback(f"[{name}] 已扫描 {scanned} 条，已删除 {deleted} 条")

            _raise_if_stop_requested(stop_callback)
            if matched_ids:
                deleted += delete_ids_in_batches(collection, matched_ids)

            total_deleted += deleted
            if progress_callback:
                progress_callback(f"[{name}] clean 集合违禁词过滤完成，扫描 {scanned} 条，删除 {deleted} 条")

        invalidate_collection_cache()
        return {"collections": len(collections), "deleted": total_deleted, "stopped": False}
    except CleanupStopRequested:
        return {"collections": len(collections), "deleted": total_deleted, "stopped": True}
    finally:
        client.close()


def run_domain_category_match(
    domain,
    category_str,
    selected_collection=None,
    progress_callback=None,
    stop_callback=None,
):
    """根据域名和分类匹配商品并导出Excel

    匹配规则(多策略):
      阶段1 — 二级分类:
        策略A: 关键词重叠评分(核心词命中>=2,按命中率打分)
        策略B: token_set_ratio >= 85 (容错词序)
        策略C: token_sort_ratio >= 95 (严格匹配)
        策略D: partial_ratio >= 75 (子串匹配,仅限长文本)
        以上对"分类"字段和"标题"字段分别执行,取最高分
      阶段2 — 一级分类(兜底): 当阶段1结果<50000时触发, 80%阈值 + 关键词重叠
      每个分类最多保留 3000 条,按匹配度评分从高到低选取
      导出上限: PER_CATEGORY_LIMIT × 分类数

    Args:
        domain: 域名(用于输出文件名)
        category_str: "一级分类|||二级分类"
        selected_collection: 集合名或"__all__"
        progress_callback: 进度回调
        stop_callback: 停止检查回调

    Returns:
        dict: {total, category_counts, file_path, stopped}
    """
    try:
        from rapidfuzz import fuzz
    except ImportError:
        if progress_callback:
            progress_callback("缺少 rapidfuzz 库，请执行: pip install rapidfuzz")
        return {"total": 0, "category_counts": {}, "file_path": "", "stopped": False}

    lines = [line.strip() for line in (category_str or "").split("\n") if line.strip()]
    categories = []
    for line in lines:
        if "|||" not in line:
            if progress_callback:
                progress_callback(f"跳过无效分类行: {line}")
            continue
        primary_cat, secondary_cat = [s.strip() for s in line.split("|||", 1)]
        if not primary_cat or not secondary_cat:
            if progress_callback:
                progress_callback(f"跳过无效分类行(一级/二级为空): {line}")
            continue
        categories.append((primary_cat, secondary_cat, line))

    if not categories:
        raise ValueError("没有有效的分类行，格式应为: 一级分类|||二级分类（每行一个）")

    PHASE1_THRESHOLD_TOKEN_SET = 85   # token_set_ratio 更宽松，处理词序不同
    PHASE1_THRESHOLD_TOKEN_SORT = 95  # token_sort_ratio 严格匹配
    PHASE1_THRESHOLD_PARTIAL = 75     # partial_ratio 子串匹配（用于标题等长文本）
    PHASE2_THRESHOLD = 80
    PER_CATEGORY_LIMIT = 2000         # 每个分类最多保留N条匹配度最高的商品
    FALLBACK_MIN = 50000
    SMALL_CATEGORY_WARN = 10
    EXPORT_DIR = "Data/商品导出"

    # ── 预处理器: 提取核心单词用于快速关键词匹配 ──
    def extract_keywords(text: str) -> set:
        words = re.findall(r"[a-z]+", text.lower())
        return {w for w in words if len(w) >= 3}

    # ── 多策略匹配(返回最高分,0表示不匹配) ──
    def multi_strategy_score(target: str, source: str) -> float:
        """多策略匹配, 返回最佳分数(0-100), 0=不匹配"""
        if not target or not source:
            return 0.0
        # 优先用 token_set_ratio (容错词序)
        set_score = fuzz.token_set_ratio(target, source)
        if set_score >= PHASE1_THRESHOLD_TOKEN_SET:
            return set_score
        # token_sort_ratio (严格)
        sort_score = fuzz.token_sort_ratio(target, source)
        if sort_score >= PHASE1_THRESHOLD_TOKEN_SORT:
            return sort_score
        # partial_ratio (子串匹配,适合长文本)
        if len(target) >= 5:
            partial_score = fuzz.partial_ratio(target, source)
            if partial_score >= PHASE1_THRESHOLD_PARTIAL:
                return partial_score
        return 0.0

    # ── 关键词重叠评分 ──
    def keyword_overlap_score(target_keywords: set, source: str) -> float:
        """根据关键词命中数量评分, 0=不匹配"""
        if not target_keywords or not source:
            return 0.0
        src_lower = source.lower()
        hits = sum(1 for kw in target_keywords if kw in src_lower)
        if hits >= 2 or (len(target_keywords) <= 3 and hits >= 1):
            # 分值与命中率正相关, 上限100
            return min(100.0, (hits / max(len(target_keywords), 1)) * 100 + 50)
        return 0.0

    # ── 单条商品匹配评分(针对某个分类) ──
    def calc_match_score(target: str, keywords: set, cat_value: str, title_value: str) -> float:
        """对一条商品分别检查分类和标题,返回最高分"""
        best = 0.0
        # 分类字段: 关键词优先(快), 模糊匹配兜底
        kw_score = keyword_overlap_score(keywords, cat_value)
        if kw_score > best:
            best = kw_score
        fuzzy_score = multi_strategy_score(target, cat_value)
        if fuzzy_score > best:
            best = fuzzy_score
        # 标题字段
        if len(title_value) >= 10:
            kw_score = keyword_overlap_score(keywords, title_value)
            if kw_score > best:
                best = kw_score
            fuzzy_score = multi_strategy_score(target, title_value)
            if fuzzy_score > best:
                best = fuzzy_score
        return best

    client = get_client()
    try:
        db = get_source_db(client)
        collections = resolve_collections(db, selected_collection)

        # 按分类存放匹配结果: {label: [(score, doc), ...]}
        matched_by_cat: Dict[str, List[Tuple[float, dict]]] = {}
        matched: List[Tuple[dict, str]] = []  # 最终展平结果, 提前初始化用于异常处理
        seen_keys = set()
        total_phase1 = 0
        total_phase2 = 0

        # ── 为每个分类预计算关键词 ──
        category_keywords = {
            sec: extract_keywords(sec) for _, sec, _ in categories
        }

        # ── 阶段1: 所有分类的二级分类匹配 ──
        for primary_cat, secondary_cat, full_cat in categories:
            if progress_callback:
                progress_callback(f"===== 处理分类: {full_cat} =====")

            sec_keywords = category_keywords[secondary_cat]
            if full_cat not in matched_by_cat:
                matched_by_cat[full_cat] = []

            for coll_name in collections:
                _raise_if_stop_requested(stop_callback)
                if not coll_name.endswith("_clean"):
                    continue

                coll = db[coll_name]
                total_in_coll = coll.estimated_document_count()
                if total_in_coll == 0:
                    continue

                if progress_callback:
                    progress_callback(f"[{full_cat}][阶段1] 扫描集合 {coll_name} ({total_in_coll}条)")

                cat_list = matched_by_cat[full_cat]
                cursor = coll.find({}, batch_size=5000)
                for doc in cursor:
                    if len(cat_list) >= PER_CATEGORY_LIMIT:
                        break

                    key = doc.get("unique_key") or str(doc["_id"])
                    if key in seen_keys:
                        continue

                    cat_value = (doc.get(CATEGORY_FIELD) or "").strip()
                    title_value = (doc.get(TITLE_FIELD) or "").strip()

                    score = calc_match_score(secondary_cat, sec_keywords, cat_value, title_value)
                    if score > 0:
                        cat_list.append((score, doc))
                        seen_keys.add(key)
                        total_phase1 += 1

                if progress_callback:
                    progress_callback(f"[{full_cat}][阶段1] {coll_name} 完成，当前该分类已匹配 {len(cat_list)} 条")

            if progress_callback:
                progress_callback(f"[{full_cat}] 阶段1完成，匹配 {len(matched_by_cat[full_cat])} 条")

        # ── 阶段2: 一级分类兜底 (当所有分类阶段1合计 < 50000) ──
        total_phase1_all = sum(len(v) for v in matched_by_cat.values())
        if total_phase1_all < FALLBACK_MIN:
            if progress_callback:
                progress_callback(f"阶段1合计 {total_phase1_all} 条 < {FALLBACK_MIN}，启用一级分类兜底(阈值{PHASE2_THRESHOLD}%)")

            for primary_cat, secondary_cat, full_cat in categories:
                # 阶段2用一级分类名匹配，存入对应分类下
                pri_keywords = extract_keywords(primary_cat)
                cat_list = matched_by_cat[full_cat]

                for coll_name in collections:
                    _raise_if_stop_requested(stop_callback)
                    if not coll_name.endswith("_clean"):
                        continue

                    coll = db[coll_name]
                    total_in_coll = coll.estimated_document_count()
                    if total_in_coll == 0:
                        continue

                    if progress_callback:
                        progress_callback(f"[{full_cat}][阶段2] 扫描集合 {coll_name}")

                    cursor = coll.find({}, batch_size=5000)
                    for doc in cursor:
                        if len(cat_list) >= PER_CATEGORY_LIMIT:
                            break

                        key = doc.get("unique_key") or str(doc["_id"])
                        if key in seen_keys:
                            continue

                        cat_value = (doc.get(CATEGORY_FIELD) or "").strip()
                        title_value = (doc.get(TITLE_FIELD) or "").strip()

                        score = calc_match_score(primary_cat, pri_keywords, cat_value, title_value)
                        if score > 0:
                            cat_list.append((score, doc))
                            seen_keys.add(key)
                            total_phase2 += 1

                    if progress_callback:
                        progress_callback(f"[{full_cat}][阶段2] {coll_name} 完成，该分类当前 {len(cat_list)} 条")

            if progress_callback:
                progress_callback(f"一级分类兜底完成，阶段2匹配 {total_phase2} 条")

        # ── 每个分类按评分降序取前 N 条 ──
        matched = []  # [(doc, label)]
        for label, items in matched_by_cat.items():
            items.sort(key=lambda x: x[0], reverse=True)
            kept = items[:PER_CATEGORY_LIMIT]
            matched.extend((doc, label) for score, doc in kept)
            if progress_callback and len(items) > PER_CATEGORY_LIMIT:
                progress_callback(f"[{label}] 排序筛选: {len(items)} → {len(kept)} 条 (保留最高分)")

        # ── 统计 ──
        category_counts = {}
        for _, label in matched:
            category_counts[label] = category_counts.get(label, 0) + 1

        if progress_callback:
            progress_callback(f"共匹配到 {len(matched)} 条商品")
            for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
                progress_callback(f"  {cat}: {count}条")
                if count < SMALL_CATEGORY_WARN:
                    progress_callback(f"  ⚠ {cat}: 数量少于{SMALL_CATEGORY_WARN}条 ({count})")

        if not matched:
            if progress_callback:
                progress_callback("未匹配到任何商品")
            return {"total": 0, "category_counts": {}, "file_path": "", "stopped": False}

        # ── 导出Excel (与product_exporter.py数据格式一致,仅修改"分类"字段) ──
        import pandas as pd
        import os
        from datetime import datetime

        EXPORT_COLUMNS = [
            "SKU", "标题", "描述", "子描述", "图片",
            "原价", "折扣价", "变体名", "变体值", "分类",
        ]

        rows_data = []
        for doc, label in matched:
            row = {}
            for field in EXPORT_COLUMNS:
                if field == "分类":
                    row[field] = label
                elif field == "SKU":
                    # 优先SKU字段, 兜底用 unique_key 或 _id
                    sku = doc.get("SKU") or doc.get("sku") or ""
                    if not sku:
                        sku = doc.get("unique_key", "")
                    if not sku:
                        sku = str(doc.get("_id", ""))
                    row[field] = str(sku)
                else:
                    val = doc.get(field)
                    if val is None:
                        val = ""
                    elif isinstance(val, list):
                        val = ", ".join(str(v).strip() for v in val if str(v).strip())
                    else:
                        val = str(val).strip()
                    row[field] = val
            rows_data.append(row)

        df = pd.DataFrame(rows_data, columns=EXPORT_COLUMNS)

        safe_domain = re.sub(r'[\\/:*?"<>|]', "_", domain) if domain else "unknown"
        os.makedirs(EXPORT_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{safe_domain}_匹配结果_{timestamp}.xlsx"
        file_path = os.path.join(EXPORT_DIR, file_name)

        df.to_excel(file_path, index=False)

        if progress_callback:
            progress_callback(f"导出文件: {file_path}")

        return {
            "total": len(matched),
            "category_counts": dict(category_counts),
            "file_path": file_path,
            "stopped": False,
        }
    except CleanupStopRequested:
        return {"total": len(matched) if matched else 0, "category_counts": {}, "file_path": "", "stopped": True}
    finally:
        client.close()


def run_category_cleanup(selected_collection=None, progress_callback=None, threshold=CATEGORY_SMALL_THRESHOLD, stop_callback=None):
    client = get_client()
    try:
        db = get_source_db(client)
        ensure_source_indexes(db)
        collections = resolve_collections(db, selected_collection)
        total_deleted = 0
        total_normalized = 0
        total_merged = 0

        for name in collections:
            _raise_if_stop_requested(stop_callback)
            collection = db[name]
            delete_ids = []
            normalize_ops = []
            deleted_chinese = 0

            category_counter = Counter()
            representative_buckets = defaultdict(Counter)

            if progress_callback:
                progress_callback(f"[{name}] 开始分类清洗与合并（单次扫描）")

            # 单次扫描：清洗 + 频率统计同时完成
            for doc in collection.find({}, {CATEGORY_FIELD: 1}, batch_size=BULK_SCAN_BATCH):
                _raise_if_stop_requested(stop_callback)
                original = text_value(doc.get(CATEGORY_FIELD))
                cleaned = clean_category(original)
                if cleaned == "DROP_ROW":
                    delete_ids.append(doc["_id"])
                    deleted_chinese += 1
                    continue
                if cleaned != original:
                    normalize_ops.append(UpdateOne({"_id": doc["_id"]}, {"$set": {CATEGORY_FIELD: cleaned}}))

                category = cleaned or "Other"
                norm_key = normalize_category_key(category)
                if not norm_key:
                    norm_key = "other"
                category_counter[norm_key] += 1
                representative_buckets[norm_key][category] += 1

            deleted = delete_ids_in_batches(collection, delete_ids) if delete_ids else 0
            normalized = bulk_write_in_batches(collection, normalize_ops) if normalize_ops else 0

            representative_map = {
                norm_key: counter.most_common(1)[0][0]
                for norm_key, counter in representative_buckets.items()
            }
            large_norm_keys = [
                norm_key
                for norm_key, count in category_counter.items()
                if count >= threshold and representative_map.get(norm_key) != "Other"
            ]
            large_categories = [representative_map[norm_key] for norm_key in large_norm_keys]

            merge_target_map = {}
            for norm_key, count in category_counter.items():
                if count >= threshold:
                    continue
                source_category = representative_map.get(norm_key, "Other")
                if source_category == "Other":
                    continue
                target_category = get_best_category_match(source_category, large_categories)
                if target_category and target_category not in {"Other", source_category}:
                    merge_target_map[norm_key] = target_category

            if merge_target_map:
                merge_ops = []
                merge_source_cats = set()
                for norm_key in merge_target_map:
                    for orig_cat in representative_buckets.get(norm_key, {}):
                        merge_source_cats.add(orig_cat)
                        trimmed = orig_cat.strip()
                        if trimmed != orig_cat:
                            merge_source_cats.add(trimmed)
                if merge_source_cats:
                    for doc in collection.find(
                        {CATEGORY_FIELD: {"$in": list(merge_source_cats)}},
                        {CATEGORY_FIELD: 1},
                        batch_size=BULK_SCAN_BATCH,
                    ):
                        _raise_if_stop_requested(stop_callback)
                        category = text_value(doc.get(CATEGORY_FIELD)) or "Other"
                        norm_key = normalize_category_key(category)
                        target = merge_target_map.get(norm_key)
                        if target and target != category:
                            merge_ops.append(UpdateOne({"_id": doc["_id"]}, {"$set": {CATEGORY_FIELD: target}}))
                merged = bulk_write_in_batches(collection, merge_ops) if merge_ops else 0
            else:
                merged = 0

            total_deleted += deleted
            total_normalized += normalized
            total_merged += merged

            if progress_callback:
                progress_callback(
                    f"[{name}] 分类清洗完成，删除主要中文分类 {deleted_chinese} 条"
                    f" | 标准化更新 {normalized} 条"
                    f" | 小分类合并 {merged} 条"
                )

        invalidate_collection_cache()
        return {
            "collections": len(collections),
            "deleted": total_deleted,
            "normalized": total_normalized,
            "merged": total_merged,
            "stopped": False,
        }
    except CleanupStopRequested:
        return {
            "collections": len(collections),
            "deleted": total_deleted,
            "normalized": total_normalized,
            "merged": total_merged,
            "stopped": True,
        }
    finally:
        client.close()


def process_excel_category_merge(file_path, threshold=CATEGORY_SMALL_THRESHOLD, progress_callback=None, stop_callback=None):
    from openpyxl import load_workbook

    try:
        if not file_path or not os.path.exists(file_path):
            raise FileNotFoundError("Excel 文件不存在")

        _raise_if_stop_requested(stop_callback)
        workbook = load_workbook(file_path)
        sheet = workbook.active

        headers = [text_value(cell.value) for cell in sheet[1]]
        normalized_headers = [header.replace("\ufeff", "").strip() for header in headers]
        category_aliases = {CATEGORY_FIELD, "category", "Category", "categories", "Categories"}
        matched_category = next((header for header in normalized_headers if header in category_aliases), "")
        if not matched_category:
            raise ValueError(f"Excel 中缺少 `{CATEGORY_FIELD}` 列")
        category_col_idx = normalized_headers.index(matched_category) + 1

        total_rows = max(sheet.max_row - 1, 0)
        if total_rows == 0:
            os.makedirs(EXPORT_DIR, exist_ok=True)
            output_path = os.path.join(
                EXPORT_DIR,
                f"categories_{os.path.basename(file_path)}",
            )
            workbook.save(output_path)
            return {
                "total": 0,
                "updated": 0,
                "small_categories": 0,
                "file_path": output_path,
                "stopped": False,
            }

        category_counter = Counter()
        representative_buckets = defaultdict(Counter)
        normalized_categories = []

        for idx, row_idx in enumerate(range(2, sheet.max_row + 1), start=1):
            _raise_if_stop_requested(stop_callback)
            value = sheet.cell(row=row_idx, column=category_col_idx).value
            cleaned = clean_category(value)
            if cleaned == "DROP_ROW":
                cleaned = "Other"
            category = cleaned or "Other"
            normalized_categories.append(category)

            norm_key = normalize_category_key(category)
            if not norm_key:
                norm_key = "other"
            category_counter[norm_key] += 1
            representative_buckets[norm_key][category] += 1

            if progress_callback and idx % 1000 == 0:
                progress_callback(f"[Excel分类处理] 统计分类中 {idx}/{total_rows}")

        public_target_map = {}
        for norm_key, count in category_counter.items():
            if count >= threshold:
                continue
            public_target_map[norm_key] = random.choice(PUBLIC_CATEGORY_CHOICES)

        updated = 0
        small_categories = len(public_target_map)
        for idx, (row_idx, category) in enumerate(zip(range(2, sheet.max_row + 1), normalized_categories), start=1):
            _raise_if_stop_requested(stop_callback)
            norm_key = normalize_category_key(category) or "other"
            target = public_target_map.get(norm_key)
            if target and target != category:
                sheet.cell(row=row_idx, column=category_col_idx).value = target
                updated += 1
            else:
                sheet.cell(row=row_idx, column=category_col_idx).value = category

            if progress_callback and idx % 1000 == 0:
                progress_callback(f"[Excel分类处理] 写入新分类中 {idx}/{total_rows}")

        os.makedirs(EXPORT_DIR, exist_ok=True)
        output_path = os.path.join(
            EXPORT_DIR,
            f"categories_{os.path.basename(file_path)}",
        )
        workbook.save(output_path)

        if progress_callback:
            progress_callback(
                f"[Excel分类处理] 完成: 总行数 {total_rows} | 小分类种类 {small_categories} | 更新分类 {updated} | 输出 {output_path}"
            )

        return {
            "total": total_rows,
            "updated": updated,
            "small_categories": small_categories,
            "file_path": output_path,
            "stopped": False,
        }
    except CleanupStopRequested:
        return {"total": 0, "updated": 0, "small_categories": 0, "file_path": "", "stopped": True}



def run_bulk_cleanup(selected_collection=None, progress_callback=None, stop_callback=None):
    client = get_client()
    try:
        db = get_source_db(client)
        ensure_source_indexes(db)
        recycle_collection = client[RECYCLE_DB][RECYCLE_COLLECTION]
        collections = resolve_collections(db, selected_collection)
        total_basic = 0
        total_image = 0
        total_english = 0
        total_forbidden = 0

        non_latin_pattern = _NON_LATIN_RE.pattern
        chinese_prohibited_pattern = (
            _CHINESE_PROHIBITED_REGEX.pattern if _CHINESE_PROHIBITED_REGEX else None
        )
        eng_prohibited_pattern = (
            _ENGLISH_PROHIBITED_REGEX.pattern if _ENGLISH_PROHIBITED_REGEX else None
        )

        projection = {TITLE_FIELD: 1, DESC_FIELD: 1, CATEGORY_FIELD: 1, IMAGE_FIELD: 1,
                      "source_url": 1, "source_category": 1}
        for field in PRICE_FIELDS:
            projection[field] = 1

        for name in collections:
            _raise_if_stop_requested(stop_callback)
            collection = db[name]

            if progress_callback:
                est = collection.estimated_document_count()
                progress_callback(f"[{name}] 全服务端批量清洗开始，{est} 条")

            # ========== Phase A: Server-side delete_many (zero Python) ==========

            r = collection.delete_many({"$or": [
                {TITLE_FIELD: None}, {TITLE_FIELD: ""},
                {DESC_FIELD: None}, {DESC_FIELD: ""},
            ]})
            total_basic += r.deleted_count
            if r.deleted_count and progress_callback:
                progress_callback(f"[{name}] 服务端: 清空标题/描述 {r.deleted_count} 条")

            r = collection.delete_many({"$expr": {
                "$lt": [{"$strLenCP": {"$ifNull": [f"${TITLE_FIELD}", ""]}}, MIN_TITLE_LEN]
            }})
            total_basic += r.deleted_count
            if r.deleted_count and progress_callback:
                progress_callback(f"[{name}] 服务端: 短标题 {r.deleted_count} 条")

            r = collection.delete_many({"$expr": {
                "$lt": [{"$strLenCP": {"$ifNull": [f"${DESC_FIELD}", ""]}}, MIN_DESC_LEN]
            }})
            total_basic += r.deleted_count
            if r.deleted_count and progress_callback:
                progress_callback(f"[{name}] 服务端: 短描述 {r.deleted_count} 条")

            r = collection.delete_many({TITLE_FIELD: {"$regex": r"^\s*\d+\s*$"}})
            total_basic += r.deleted_count
            if r.deleted_count and progress_callback:
                progress_callback(f"[{name}] 服务端: 数字标题 {r.deleted_count} 条")

            r = collection.delete_many({"$or": [
                {IMAGE_FIELD: {"$in": [None, "", []]}},
                {IMAGE_FIELD: {"$regex": r"coming-soon|noimage|default|\.svg", "$options": "i"}},
            ]})
            total_image += r.deleted_count
            if r.deleted_count and progress_callback:
                progress_callback(f"[{name}] 服务端: 异常图片 {r.deleted_count} 条")

            r = collection.delete_many({TITLE_FIELD: {"$regex": non_latin_pattern}})
            total_english += r.deleted_count
            if r.deleted_count and progress_callback:
                progress_callback(f"[{name}] 服务端: 非拉丁标题 {r.deleted_count} 条")

            r = collection.delete_many({TITLE_FIELD: {"$regex": _ACCENT_REGEX}})
            total_english += r.deleted_count
            if r.deleted_count and progress_callback:
                progress_callback(f"[{name}] 服务端: 欧洲重音 {r.deleted_count} 条")

            if progress_callback:
                progress_callback(f"[{name}] 服务端 A 阶段完成，待处理: basic={total_basic} image={total_image} english={total_english}")

            # ========== Phase B: Price + ratio in single Python cursor ==========
            if progress_callback:
                progress_callback(f"[{name}] 扫描价格和拉丁比率 ...")
            est = collection.estimated_document_count()
            basic_ids = []
            english_ids = []
            scanned = 0
            cursor = collection.find({}, {
                "_id": 1, TITLE_FIELD: 1, CATEGORY_FIELD: 1, DESC_FIELD: 1,
                **{pf: 1 for pf in PRICE_FIELDS}
            }, batch_size=5000, no_cursor_timeout=True)
            try:
                for doc in cursor:
                    _raise_if_stop_requested(stop_callback)
                    scanned += 1
                    if scanned % 5000 == 0 and progress_callback:
                        progress_callback(f"[{name}] 扫描中 {scanned}/{est} ...")

                    bad_price = False
                    for pf in PRICE_FIELDS:
                        val = doc.get(pf)
                        if val is not None and val != "":
                            try:
                                fval = float(val)
                                if fval < MIN_PRICE or fval > MAX_PRICE:
                                    basic_ids.append(doc["_id"])
                                    bad_price = True
                                    break
                            except (ValueError, TypeError):
                                pass
                    if bad_price:
                        continue

                    title = doc.get(TITLE_FIELD) or ""
                    category = doc.get(CATEGORY_FIELD) or ""
                    desc = doc.get(DESC_FIELD) or ""
                    prime_text = (title + " " + category).strip()
                    comb_text = (prime_text + " " + desc).strip()
                    pl = len(_LETTER_RE.findall(prime_text))
                    if pl == 0:
                        continue
                    pL = len(_LATIN_RE.findall(prime_text))
                    cl = len(_LETTER_RE.findall(comb_text))
                    cL = len(_LATIN_RE.findall(comb_text))
                    dNonLat = bool(_NON_LATIN_RE.search(desc))
                    ratio_match = False
                    if pl >= 12 and (pL / pl) < 0.55:
                        ratio_match = True
                    elif cl >= 80 and (cL / cl) < 0.45 and pL < 8:
                        ratio_match = True
                    elif dNonLat:
                        if pl < 8:
                            ratio_match = True
                        elif pL / pl < 0.5:
                            ratio_match = True
                    if ratio_match:
                        english_ids.append(doc["_id"])
            finally:
                cursor.close()

            for i in range(0, len(basic_ids), DELETE_BATCH_SIZE):
                batch = basic_ids[i:i + DELETE_BATCH_SIZE]
                total_basic += collection.delete_many({"_id": {"$in": batch}}).deleted_count
            if basic_ids and progress_callback:
                progress_callback(f"[{name}] 异常价格 {len(basic_ids)} 条")

            for i in range(0, len(english_ids), DELETE_BATCH_SIZE):
                batch = english_ids[i:i + DELETE_BATCH_SIZE]
                total_english += collection.delete_many({"_id": {"$in": batch}}).deleted_count
            if english_ids and progress_callback:
                progress_callback(f"[{name}] 拉丁比率 {len(english_ids)} 条")

            # ========== Phase C: Prohibited keywords ==========
            prohibited_ids = []

            if _AC_AUTOMATON:
                if progress_callback:
                    progress_callback(f"[{name}] 违禁词扫描开始（Aho-Corasick 全量扫描）")
                est = collection.estimated_document_count()
                scanned = 0
                cursor = collection.find({}, projection, batch_size=FORBIDDEN_BATCH_SIZE, no_cursor_timeout=True)
                try:
                    for doc in cursor:
                        _raise_if_stop_requested(stop_callback)
                        scanned += 1
                        if scanned % 5000 == 0 and progress_callback:
                            progress_callback(f"[{name}] 违禁词扫描 {scanned}/{est} ...")
                        match = find_prohibited_match(doc)
                        if not match:
                            continue
                        match_field, match_keyword = match
                        recycle_doc = dict(doc)
                        oid = recycle_doc.pop("_id", None)
                        recycle_doc["original_id"] = str(oid) if oid is not None else ""
                        recycle_doc["recycle_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        recycle_doc["recycle_reason"] = "apparel_category" if str(match_keyword).startswith("apparel:") else "forbidden_keyword"
                        recycle_doc["recycle_source_collection"] = name
                        recycle_doc["recycle_match_field"] = match_field
                        recycle_doc["recycle_match_keyword"] = match_keyword
                        prohibited_ids.append((recycle_doc, doc["_id"]))
                finally:
                    cursor.close()
                if progress_callback and scanned:
                    progress_callback(f"[{name}] 违禁词扫描完成，已扫描 {scanned} 条 / 命中 {len(prohibited_ids)} 条")
            else:
                combined_or = []
                if chinese_prohibited_pattern:
                    for field in (TITLE_FIELD, CATEGORY_FIELD, DESC_FIELD):
                        combined_or.append({field: {"$regex": chinese_prohibited_pattern, "$options": "i"}})
                if eng_prohibited_pattern:
                    for field in (TITLE_FIELD, CATEGORY_FIELD, DESC_FIELD):
                        combined_or.append({field: {"$regex": eng_prohibited_pattern, "$options": "i"}})
                if combined_or:
                    if progress_callback:
                        progress_callback(f"[{name}] 违禁词匹配阶段开始")
                    scanned = 0
                    for doc in collection.find({"$or": combined_or}, projection, batch_size=FORBIDDEN_BATCH_SIZE):
                        _raise_if_stop_requested(stop_callback)
                        scanned += 1
                        if scanned % FORBIDDEN_PROGRESS_EVERY == 0 and progress_callback:
                            progress_callback(f"[{name}] 违禁词扫描 {scanned} 条 ...")
                        match = find_prohibited_match(doc)
                        if not match:
                            continue
                        match_field, match_keyword = match
                        recycle_doc = dict(doc)
                        oid = recycle_doc.pop("_id", None)
                        recycle_doc["original_id"] = str(oid) if oid is not None else ""
                        recycle_doc["recycle_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        recycle_doc["recycle_reason"] = "apparel_category" if str(match_keyword).startswith("apparel:") else "forbidden_keyword"
                        recycle_doc["recycle_source_collection"] = name
                        recycle_doc["recycle_match_field"] = match_field
                        recycle_doc["recycle_match_keyword"] = match_keyword
                        prohibited_ids.append((recycle_doc, doc["_id"]))
                    if progress_callback and scanned:
                        progress_callback(f"[{name}] 违禁词扫描完成，匹配 {scanned} 条 / 命中 {len(prohibited_ids)} 条")

            for i in range(0, len(prohibited_ids), FORBIDDEN_BATCH_SIZE):
                batch = prohibited_ids[i:i + FORBIDDEN_BATCH_SIZE]
                recycle_collection.insert_many([r for r, _ in batch], ordered=False)
                total_forbidden += collection.delete_many({"_id": {"$in": [d for _, d in batch]}}).deleted_count

            collection.update_many({}, {"$set": {"子描述": ""}})
            if total_forbidden and progress_callback:
                progress_callback(f"[{name}] 违禁词处理完成，移入回收站 {total_forbidden} 条")

            if progress_callback:
                progress_callback(
                    f"[{name}] 完成 -> basic {total_basic} / image {total_image}"
                    f" / english {total_english} / forbidden {total_forbidden}"
                )

        invalidate_collection_cache()
        return {"collections": len(collections), "basic_deleted": total_basic, "image_deleted": total_image,
                "english_deleted": total_english, "forbidden_moved": total_forbidden, "stopped": False}
    except CleanupStopRequested:
        return {"collections": len(collections), "basic_deleted": total_basic, "image_deleted": total_image,
                "english_deleted": total_english, "forbidden_moved": total_forbidden, "stopped": True}
    finally:
        client.close()


def run_bulk_cleanup_parallel(selected_collection=None, progress_callback=None, stop_callback=None, max_workers=4):
    client = get_client()
    try:
        db = get_source_db(client)
        collections = resolve_collections(db, selected_collection)
        if not collections:
            return {"collections": 0, "basic_deleted": 0, "image_deleted": 0,
                    "english_deleted": 0, "forbidden_moved": 0, "stopped": False}

        if len(collections) == 1:
            return run_bulk_cleanup(selected_collection=collections[0],
                                    progress_callback=progress_callback, stop_callback=stop_callback)

        if progress_callback:
            progress_callback(f"并行处理 {len(collections)} 个集合（{max_workers} 线程）")

        results = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            def _run_one(coll):
                local_client = get_client()
                try:
                    return run_bulk_cleanup(selected_collection=coll, stop_callback=stop_callback)
                finally:
                    local_client.close()

            futures = {executor.submit(_run_one, c): c for c in collections}
            for future in as_completed(futures):
                coll = futures[future]
                try:
                    res = future.result()
                    results.append(res)
                    if progress_callback:
                        d = res.get("basic_deleted", 0) + res.get("image_deleted", 0) + res.get("english_deleted", 0)
                        f = res.get("forbidden_moved", 0)
                        progress_callback(f"[{coll}] 完成 | 删除 {d} | 违禁 {f}")
                except CleanupStopRequested:
                    if progress_callback:
                        progress_callback(f"[{coll}] 已停止")
                    results.append({"collections": 1, "basic_deleted": 0, "image_deleted": 0,
                                    "english_deleted": 0, "forbidden_moved": 0, "stopped": True})
                except Exception as e:
                    if progress_callback:
                        progress_callback(f"[{coll}] 错误: {e}")
                    results.append({"collections": 1, "basic_deleted": 0, "image_deleted": 0,
                                    "english_deleted": 0, "forbidden_moved": 0, "stopped": True})

        merged = {"collections": len(collections), "basic_deleted": 0, "image_deleted": 0,
                  "english_deleted": 0, "forbidden_moved": 0, "stopped": False}
        for r in results:
            merged["basic_deleted"] += r.get("basic_deleted", 0)
            merged["image_deleted"] += r.get("image_deleted", 0)
            merged["english_deleted"] += r.get("english_deleted", 0)
            merged["forbidden_moved"] += r.get("forbidden_moved", 0)
            if r.get("stopped"):
                merged["stopped"] = True
        return merged
    finally:
        client.close()


def run_full_cleanup(selected_collection=None, progress_callback=None, stop_callback=None, parallel=False):
    if parallel:
        if progress_callback:
            progress_callback("开始批量清洗阶段（合并基础/图片/英文/违禁词过滤）")
        bulk = run_bulk_cleanup_parallel(
            selected_collection=selected_collection,
            progress_callback=progress_callback,
            stop_callback=stop_callback,
        )
    else:
        if progress_callback:
            progress_callback("开始批量清洗阶段（合并基础/图片/英文/违禁词过滤）")
        bulk = run_bulk_cleanup(
            selected_collection=selected_collection,
            progress_callback=progress_callback,
            stop_callback=stop_callback,
        )

    if bulk.get("stopped"):
        return {
            "collections": bulk.get("collections", 0),
            "basic_deleted": bulk.get("basic_deleted", 0),
            "image_deleted": bulk.get("image_deleted", 0),
            "english_deleted": bulk.get("english_deleted", 0),
            "forbidden_moved": bulk.get("forbidden_moved", 0),
            "category_deleted": 0,
            "category_normalized": 0,
            "category_merged": 0,
            "stopped": True,
        }

    if progress_callback:
        progress_callback("开始分类清洗与合并阶段")
    category = run_category_cleanup(
        selected_collection=selected_collection,
        progress_callback=progress_callback,
        stop_callback=stop_callback,
    )

    return {
        "collections": max(bulk.get("collections", 0), category.get("collections", 0)),
        "basic_deleted": bulk.get("basic_deleted", 0),
        "image_deleted": bulk.get("image_deleted", 0),
        "english_deleted": bulk.get("english_deleted", 0),
        "forbidden_moved": bulk.get("forbidden_moved", 0),
        "category_deleted": category.get("deleted", 0),
        "category_normalized": category.get("normalized", 0),
        "category_merged": category.get("merged", 0),
        "stopped": bool(category.get("stopped")),
    }


def run_extract_clean(selected_collection=None, progress_callback=None, stop_callback=None):
    """从 shopify_data_new 中提取好数据到 {类目}_clean，脏数据直接删除。

    对每个类目集合依次执行: basic → image → english → category 检查。
    通过的数据写入 {类目}_clean 并做小分类合并；未通过的数据直接删除。
    原类目集合中所有已处理的数据都被清除，但集合本身不删除。
    不执行违禁词过滤。
    """
    client = get_client()
    db = client[CLEAN_DB]
    collections = resolve_collections(db, selected_collection)
    results = []

    try:
        if progress_callback:
            progress_callback(describe_collection_scope(selected_collection, collections, skip_clean=True))
        if not collections:
            return {"collections": results, "stopped": False}

        for coll_name in collections:
            _raise_if_stop_requested(stop_callback)
            if coll_name.endswith("_clean"):
                continue
            source_coll = db[coll_name]
            target_coll_name = f"{coll_name}_clean"
            target_coll = db[target_coll_name]
            ensure_collection_indexes(source_coll, ("unique_key", "source_url", "crawl_time"))
            ensure_collection_indexes(target_coll, ("unique_key", CATEGORY_FIELD))

            total = source_coll.estimated_document_count()
            if total == 0:
                continue

            if progress_callback:
                progress_callback(f"[{coll_name}] 开始提取清洗: 共 {total} 条 → {target_coll_name}")

            projection = {
                TITLE_FIELD: 1, DESC_FIELD: 1, CATEGORY_FIELD: 1, IMAGE_FIELD: 1,
                "source_url": 1, "unique_key": 1, "crawl_time": 1,
            }
            for field in PRICE_FIELDS:
                projection[field] = 1

            clean_batch = []
            processed_ids_batch = []
            passed_count = 0
            failed_count = 0
            normalized_count = 0
            processed = 0
            inserted = 0
            existing_count = 0
            deleted_total = 0
            write_failed = False

            category_counter = Counter()
            representative_buckets = defaultdict(Counter)

            def flush_extract_batch():
                nonlocal clean_batch, processed_ids_batch, inserted, existing_count, deleted_total, write_failed
                if not processed_ids_batch:
                    return True
                try:
                    if clean_batch:
                        batch_inserted, batch_existing = write_clean_docs_batch(target_coll, clean_batch)
                        inserted += batch_inserted
                        existing_count += batch_existing
                    deleted_total += delete_ids_in_batches(source_coll, processed_ids_batch)
                except Exception as e:
                    write_failed = True
                    if progress_callback:
                        progress_callback(
                            f"[{coll_name}] 批量写入 {target_coll_name} 失败: {str(e)[:200]}，"
                            f"当前批次源数据未删除"
                        )
                    return False
                finally:
                    clean_batch = []
                    processed_ids_batch = []
                return True

            cursor = source_coll.find({}, projection).batch_size(BULK_SCAN_BATCH)
            for doc in cursor:
                _raise_if_stop_requested(stop_callback)
                processed += 1
                processed_ids_batch.append(doc["_id"])

                basic_reason = get_basic_delete_reason(doc)
                if basic_reason:
                    failed_count += 1
                elif has_brand_keyword(doc):
                    failed_count += 1
                elif has_bad_image(doc):
                    failed_count += 1
                else:
                    non_english_reason = get_non_english_reason(doc)
                    if non_english_reason:
                        failed_count += 1
                    else:
                        original_category = text_value(doc.get(CATEGORY_FIELD))
                        cleaned_category = clean_category(original_category)
                        if cleaned_category == "DROP_ROW":
                            failed_count += 1
                        else:
                            clean_doc = {k: v for k, v in doc.items() if k != "_id"}
                            clean_doc["子描述"] = ""
                            if cleaned_category and cleaned_category != original_category:
                                clean_doc[CATEGORY_FIELD] = cleaned_category
                            clean_batch.append(clean_doc)
                            passed_count += 1

                            category = cleaned_category or "Other"
                            norm_key = normalize_category_key(category)
                            if not norm_key:
                                norm_key = "other"
                            category_counter[norm_key] += 1
                            representative_buckets[norm_key][category] += 1

                if processed % 1000 == 0 and progress_callback:
                    progress_callback(f"[{coll_name}] 检查中 {processed}/{total} | 通过 {passed_count} 失败 {failed_count}")

                if len(processed_ids_batch) >= WRITE_BATCH_SIZE and not flush_extract_batch():
                    break

            if processed_ids_batch:
                flush_extract_batch()

            if progress_callback:
                if write_failed:
                    progress_callback(
                        f"[{coll_name}] 提取部分完成: 通过检查 {passed_count} 条，"
                        f"实际写入 {target_coll_name} {inserted} 条，"
                        f"已存在跳过 {existing_count} 条，"
                        f"未通过检查 {failed_count} 条，源集合删除 {deleted_total} 条，"
                        f"未处理/失败批次仍保留在源集合"
                    )
                else:
                    progress_callback(
                        f"[{coll_name}] 提取完成: 通过检查 {passed_count} 条，"
                        f"实际写入 {target_coll_name} {inserted} 条，"
                        f"已存在跳过 {existing_count} 条，"
                        f"不合格 {failed_count} 条已丢弃，源集合删除 {deleted_total} 条"
                    )

            merged = 0
            try:
                new_total = target_coll.estimated_document_count()
                if new_total > 0 and category_counter:
                    if progress_callback:
                        progress_callback(f"[{target_coll_name}] 开始分类合并 ({new_total} 条)")

                    representative_map = {
                        norm_key: counter.most_common(1)[0][0]
                        for norm_key, counter in representative_buckets.items()
                    }
                    large_norm_keys = [
                        norm_key
                        for norm_key, count in category_counter.items()
                        if count >= CATEGORY_SMALL_THRESHOLD and representative_map.get(norm_key) != "Other"
                    ]
                    large_categories = [representative_map[norm_key] for norm_key in large_norm_keys]

                    merge_target_map = {}
                    for norm_key, count in category_counter.items():
                        if count >= CATEGORY_SMALL_THRESHOLD:
                            continue
                        source_cat = representative_map.get(norm_key, "Other")
                        if source_cat == "Other":
                            continue
                        target_cat = get_best_category_match(source_cat, large_categories)
                        if target_cat and target_cat not in {"Other", source_cat}:
                            merge_target_map[norm_key] = target_cat

                    if merge_target_map:
                        merge_ops = []
                        merge_source_cats = set()
                        for nk in merge_target_map:
                            for orig_cat in representative_buckets.get(nk, {}):
                                merge_source_cats.add(orig_cat)
                                trimmed = orig_cat.strip()
                                if trimmed != orig_cat:
                                    merge_source_cats.add(trimmed)
                        if merge_source_cats:
                            for doc in target_coll.find(
                                {CATEGORY_FIELD: {"$in": list(merge_source_cats)}},
                                {CATEGORY_FIELD: 1},
                                batch_size=BULK_SCAN_BATCH,
                            ):
                                _raise_if_stop_requested(stop_callback)
                                cat_val = text_value(doc.get(CATEGORY_FIELD)) or "Other"
                                nk = normalize_category_key(cat_val)
                                target = merge_target_map.get(nk)
                                if target and target != cat_val:
                                    merge_ops.append(UpdateOne({"_id": doc["_id"]}, {"$set": {CATEGORY_FIELD: target}}))
                        merged = bulk_write_in_batches(target_coll, merge_ops) if merge_ops else 0

                    if progress_callback:
                        progress_callback(f"[{target_coll_name}] 分类合并完成: 合并 {merged} 条")
            except Exception as e:
                if progress_callback:
                    progress_callback(f"[{target_coll_name}] 分类合并出错: {str(e)[:200]}")

            results.append({
                "collection": coll_name,
                "total": total,
                "inserted": inserted,
                "existing": existing_count,
                "passed": passed_count,
                "failed": failed_count,
                "normalized": normalized_count,
                "merged": merged,
                "deleted": deleted_total,
                "write_failed": write_failed,
            })

        invalidate_collection_cache()
        return {"collections": results, "stopped": False}
    except CleanupStopRequested:
        return {"collections": results, "stopped": True}
    finally:
        client.close()


def run_staging_to_clean(selected_collection=None, progress_callback=None, stop_callback=None):
    """Read from source collection, run full cleaning checks, write clean data to <coll>_clean.

    Only data that passes ALL checks is written to the _clean collection.
    Forbidden-keyword matches are moved to the recycle bin.
    Failed data (basic/image/english/chinese-category) is dropped permanently.
    After migration, the source collection is dropped.
    """
    client = get_client()
    staging_db = client[STAGING_DB]
    clean_db = client[CLEAN_DB]
    recycle_db = client[RECYCLE_DB]
    collections = resolve_collections(staging_db, selected_collection)
    results = []

    try:
        for coll_name in collections:
            _raise_if_stop_requested(stop_callback)
            if coll_name.endswith("_clean"):
                continue
            staging_coll = staging_db[coll_name]
            clean_coll = clean_db[f"{coll_name}_clean"]
            recycle_coll = recycle_db[RECYCLE_COLLECTION]

            total = staging_coll.count_documents({})
            if total == 0:
                continue

            projection = {
                TITLE_FIELD: 1, DESC_FIELD: 1, CATEGORY_FIELD: 1, IMAGE_FIELD: 1,
                "source_url": 1, "unique_key": 1, "crawl_time": 1,
            }
            for field in PRICE_FIELDS:
                projection[field] = 1

            clean_batch = []
            recycle_batch = []
            failed_count = 0
            processed = 0

            cursor = staging_coll.find({}, projection).batch_size(BULK_SCAN_BATCH)

            for doc in cursor:
                _raise_if_stop_requested(stop_callback)
                processed += 1

                basic_reason = get_basic_delete_reason(doc)
                if basic_reason:
                    failed_count += 1
                    continue

                if has_bad_image(doc):
                    failed_count += 1
                    continue

                non_english_reason = get_non_english_reason(doc)
                if non_english_reason:
                    failed_count += 1
                    continue

                original_category = text_value(doc.get(CATEGORY_FIELD))
                cleaned_category = clean_category(original_category)
                if cleaned_category == "DROP_ROW":
                    failed_count += 1
                    continue

                forbidden_match = find_prohibited_match(doc)
                if forbidden_match:
                    field, keyword = forbidden_match
                    recycle_doc = {k: v for k, v in doc.items() if k != "_id"}
                    recycle_doc["original_id"] = str(doc["_id"])
                    recycle_doc["recycle_time"] = datetime.now()
                    recycle_doc["recycle_reason"] = "forbidden_keyword"
                    recycle_doc["recycle_source_collection"] = coll_name
                    recycle_doc["recycle_match_field"] = field
                    recycle_doc["recycle_match_keyword"] = keyword
                    recycle_batch.append(recycle_doc)
                    continue

                clean_doc = {k: v for k, v in doc.items() if k != "_id"}
                clean_doc["子描述"] = ""
                if cleaned_category and cleaned_category != original_category:
                    clean_doc[CATEGORY_FIELD] = cleaned_category
                clean_batch.append(clean_doc)

                if processed % 1000 == 0 and progress_callback:
                    progress_callback(f"[{coll_name}] 检查中 {processed}/{total}")

            inserted = 0
            recycled = 0

            if clean_batch:
                try:
                    existing_keys = set()
                    candidate_keys = [d.get("unique_key", "") for d in clean_batch if d.get("unique_key")]
                    if candidate_keys:
                        for existing in clean_coll.find({"unique_key": {"$in": candidate_keys}}, {"unique_key": 1}):
                            existing_keys.add(existing.get("unique_key", ""))
                    to_insert = [d for d in clean_batch if not d.get("unique_key") or d["unique_key"] not in existing_keys]
                    if to_insert:
                        clean_coll.insert_many(to_insert, ordered=False)
                        inserted = len(to_insert)
                except Exception as e:
                    if progress_callback:
                        progress_callback(f"[{coll_name}] 写入清洗库失败: {str(e)[:200]}")

            if recycle_batch:
                try:
                    recycle_coll.insert_many(recycle_batch, ordered=False)
                    recycled = len(recycle_batch)
                except Exception as e:
                    if progress_callback:
                        progress_callback(f"[{coll_name}] 写入回收站失败: {str(e)[:200]}")

            staging_coll.drop()

            results.append({
                "collection": coll_name,
                "total": total,
                "inserted": inserted,
                "recycled": recycled,
                "failed": failed_count,
            })

            if progress_callback:
                progress_callback(
                    f"[{coll_name}] 完成: 共 {total}, "
                    f"入库 {inserted}, 违禁回收 {recycled}, 不合格 {failed_count}"
                )

        invalidate_collection_cache()
        return {"collections": results, "stopped": False}
    except CleanupStopRequested:
        return {"collections": results, "stopped": True}
    finally:
        client.close()
