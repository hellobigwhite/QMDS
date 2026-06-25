import os
from openpyxl import load_workbook, Workbook
import xlwings as xw
import random
import string
import time


def add_suffix_to_domain_column(workbook, suffix_mode="none", custom_suffix=None, part_index=None):
    """
    给 Excel 表格中的“原站域名”列添加后缀。
    
    :param workbook: 已加载的工作簿对象
    :param suffix_mode: 后缀模式 ('none', 'custom', 'part')
    :param custom_suffix: 自定义后缀内容
    :param part_index: 当前的分卷索引
    """
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    domain_column_index = header.index("原站域名")  # 假设“原站域名”在第一行
    
    rows = list(sheet.iter_rows(min_row=2))  # 将生成器转换为列表
    
    for row in rows:
        domain_value = row[domain_column_index].value
        if suffix_mode == "custom" and custom_suffix:
            row[domain_column_index].value = f"{domain_value}_{custom_suffix}"
        elif suffix_mode == "part" and part_index is not None:
            # 为每个分卷添加后缀
            row[domain_column_index].value = f"{domain_value}_part{part_index}"
    
    # print(f"文件 {workbook} 后缀已添加！")


def split_excel_file(file_path, rows_per_file=5000, suffix_mode="none", custom_suffix=None):
    """
    将指定的 Excel 文件按行数拆分为多个文件，并在“原站域名”列添加后缀。
    
    :param file_path: 原始 Excel 文件路径
    :param rows_per_file: 每个拆分文件的最大行数（包含标题行）
    :param suffix_mode: 后缀模式 ('none', 'custom', 'part')
    :param custom_suffix: 自定义后缀内容
    """
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    base_file_name = os.path.splitext(os.path.basename(file_path))[0]
    output_folder = os.path.join(os.path.dirname(file_path), f"{base_file_name}_split")
    os.makedirs(output_folder, exist_ok=True)

    row_buffer = []
    part_index = 1
    row_count = 0

    print(f"正在处理文件: {file_path}...")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_buffer.append(row)
        row_count += 1

        if row_count == rows_per_file:
            save_split_file(output_folder, base_file_name, part_index, header, row_buffer, suffix_mode, custom_suffix)
            row_buffer = []
            row_count = 0
            part_index += 1

    if row_buffer:
        save_split_file(output_folder, base_file_name, part_index, header, row_buffer, suffix_mode, custom_suffix)

    print(f"文件 {file_path} 已完成拆分，结果保存在: {output_folder}")
    return output_folder


def save_split_file(output_folder, base_name, part_index, header, data, suffix_mode, custom_suffix):
    """
    保存拆分后的数据为新文件，并在文件名后添加随机 2 个字母 + 时间戳以保证唯一性。
    
    :param output_folder: 输出文件夹路径
    :param base_name: 原始文件名（不带扩展名）
    :param part_index: 当前拆分文件索引
    :param header: 表头
    :param data: 数据行
    :param suffix_mode: 后缀模式 ('none', 'custom', 'part')
    :param custom_suffix: 自定义后缀内容
    """
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.append(header)
    for row in data:
        new_sheet.append(row)

    # 添加后缀到“原站域名”列
    add_suffix_to_domain_column(new_workbook, suffix_mode, custom_suffix, part_index)

    # 生成随机 2 个字母
    random_letters = ''.join(random.choices(string.ascii_uppercase, k=2))
    # 获取当前时间戳
    timestamp = int(time.time())

    # 生成唯一文件名
    output_file = os.path.join(output_folder, f"{base_name}_part{part_index}_{random_letters}{timestamp}.xlsx")
    
    new_workbook.save(output_file)
    print(f"已生成文件: {output_file}")


def save_with_xlwings(folder_path: str) -> None:
    """
    使用 Excel 应用程序重新保存文件夹中的 .xlsx 文件，以解决兼容性问题。
    
    :param folder_path: 包含 .xlsx 文件的文件夹路径
    """
    if not os.path.isdir(folder_path):
        print("路径无效，请提供有效的文件夹路径。")
        return

    for file_name in os.listdir(folder_path):
        if file_name.lower().endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            print(f"正在处理文件: {file_name}...")

            try:
                app = xw.App(visible=False)
                app.display_alerts = False
                workbook = app.books.open(file_path)
                workbook.save(file_path)
                print(f"文件 {file_name} 已成功修复。")
                workbook.close()
            except Exception as e:
                print(f"处理文件 {file_name} 时出错：{e}")
                with open("error_log.txt", "a") as log_file:
                    log_file.write(f"文件 {file_name} 处理失败：{e}\n")
            finally:
                app.quit()


def run(file_path, rows_per_file=5000, suffix_mode="none", custom_suffix=None):
    """
    处理文件路径字符串，并确保传递的是路径，而不是 Workbook 对象。
    """
    if isinstance(file_path, str):  # 确保传递的是文件路径字符串
        output_folder = split_excel_file(file_path, rows_per_file, suffix_mode, custom_suffix)
        save_with_xlwings(output_folder)
    else:
        print("传入了无效的文件路径：", file_path)
