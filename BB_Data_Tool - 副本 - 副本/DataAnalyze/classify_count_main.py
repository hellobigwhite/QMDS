from openpyxl import load_workbook, Workbook
from collections import defaultdict
import os

def save_category_counts(file_path, category_counts):
    """
    将统计的 Categories 计数结果保存到新的 Excel 和 Markdown 文件中（按数量从大到小排序）。
    """
    output_folder = os.path.join(os.path.dirname(file_path), "数据分析")
    os.makedirs(output_folder, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(file_path))[0]
    
    # 按照计数从大到小排序
    sorted_counts = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)

    # 保存为 Excel 文件
    excel_file = os.path.join(output_folder, f"{base_name}_category_count.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["Category", "Count"])
    
    for category, count in sorted_counts:
        ws.append([category, count])
    
    wb.save(excel_file)
    print(f"统计结果已保存: {excel_file}")

    # 保存为 Markdown 文件（不打印内容）
    md_file = os.path.join(output_folder, f"{base_name}_category_count.md")
    with open(md_file, "w", encoding="utf-8") as md:
        md.write(f"# Categories 统计结果\n\n")
        md.write(f"## 文件: {os.path.basename(file_path)}\n\n")
        md.write("| Category | Count |\n")
        md.write("|----------|-------|\n")
        for category, count in sorted_counts:
            md.write(f"| {category} | {count} |\n")

    print(f"Markdown 统计结果已生成: {md_file}")

def run(file_path):
    """
    统计 Excel 表格中 'Categories' 列的计数，并保存到新文件。
    """
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active

    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

    if "Categories" not in header:
        print(f"错误：文件 {file_path} 中未找到 'Categories' 列！")
        return None

    category_index = header.index("Categories")
    category_counts = defaultdict(int)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        category = row[category_index] if row[category_index] is not None else "(无分类)"
        category_counts[category] += 1

    # 生成统计结果文件（Excel 和 Markdown）
    save_category_counts(file_path, category_counts)
    return category_counts
